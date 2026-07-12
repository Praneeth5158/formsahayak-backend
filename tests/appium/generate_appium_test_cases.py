import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define the dataset of 310 unique Appium Mobile Automation Test Cases
# Format: (Module, Feature, Test Scenario, Precondition, Test Steps, Test Data, Expected Result, Priority, Severity, Automation Type, Appium Locator Strategy, Automation Candidate, Suites)
tc_list = [
    # A. App Installation Testing (15 TCs)
    (
        "App Installation", "Fresh Install", "Verify successful fresh installation of the application APK",
        "Android device is connected. No previous version of FormSahayak installed.",
        "1. Push the APK to the device.\n2. Execute installation command.\n3. Verify app icon appears on launcher.",
        "APK Path: app-debug.apk",
        "Application installs successfully. FormSahayak icon is present in the app drawer.",
        "High", "Critical", "Appium Native", "N/A (ADB install command)", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "App Installation", "Upgrade Install", "Verify application upgrade from older version to new version",
        "FormSahayak version 1.0.0 is installed with active user logged in.",
        "1. Install newer version APK using upgrade/reinstall flags.\n2. Launch application.\n3. Verify app launches without crashing and user session is retained.",
        "New APK Path: app-debug-v1.1.0.apk",
        "Application upgrades successfully. Previous session is retained, and dashboard loads directly.",
        "High", "Critical", "Appium Native", "N/A (ADB install -r command)", "Yes",
        ["Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "App Installation", "Reinstallation", "Verify clean reinstallation of the application after uninstalling",
        "FormSahayak is already installed on the device.",
        "1. Uninstall the FormSahayak application.\n2. Fresh install the application.\n3. Launch application and verify login screen is loaded.",
        "APK Path: app-debug.apk",
        "Uninstall succeeds. Reinstallation succeeds. Launch loads login screen with zero cached user data.",
        "High", "Major", "Appium Native", "accessibilityId(\"login_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "App Installation", "App Launch", "Verify app launch times and behavior on first startup",
        "Application is freshly installed on the device.",
        "1. Tap the FormSahayak app icon in the launcher.\n2. Measure time taken to show Splash screen and transition to Login/Language screen.",
        "N/A",
        "App launches within 2 seconds. Language selection or login screen is interactive.",
        "High", "Critical", "Appium Native", "accessibilityId(\"english_lang_checkbox\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "App Installation", "App Launch", "Verify app launch behavior when device has no network on first startup",
        "Application is freshly installed on the device. Device internet is disconnected.",
        "1. Disable Wi-Fi and Mobile Data.\n2. Tap the FormSahayak app icon.\n3. Observe app startup and error messages.",
        "Network State: Offline",
        "App launches, shows splash, and displays offline warning bar or dialog.",
        "Medium", "Major", "Appium Native", "id(\"com.formsahayak.app:id/offline_dialog_title\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "App Installation", "Uninstall", "Verify that uninstalling the app removes all user local storage and cache",
        "App is installed and user has uploaded several forms locally.",
        "1. Uninstall FormSahayak via device settings or ADB.\n2. Check for leftover folders under /data/data/com.formsahayak.app/.",
        "Package: com.formsahayak.app",
        "All local files, Shared Preferences, and database files are completely deleted.",
        "Medium", "Major", "Appium Native", "N/A (ADB shell check)", "No",
        ["Regression"]
    ),
    (
        "App Installation", "App Launch", "Verify app launches correctly when launched via deep link",
        "App is installed. User is not logged in.",
        "1. Send a deep link command for login page via ADB.\n2. Verify target screen loads.",
        "Deep Link: formsahayak://login",
        "App launches and immediately displays the login screen.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"login_email\")", "Yes",
        ["Regression"]
    ),
    (
        "App Installation", "App Launch", "Verify launch from background state",
        "App is running in the foreground.",
        "1. Press Home button to send app to background.\n2. Tap the app icon in the launcher or open from recents.\n3. Verify app state is restored.",
        "N/A",
        "App returns to foreground immediately at the exact screen and state it was left.",
        "High", "Major", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "App Installation", "App Launch", "Verify app launch behavior after clearing cache and force stopping",
        "App is installed and user has cached items.",
        "1. Force stop app and clear cache via ADB/Settings.\n2. Launch app.\n3. Verify app launches successfully.",
        "N/A",
        "App launches correctly. Performance might show a minor load time during initial caching.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Regression"]
    ),
    (
        "App Installation", "App Launch", "Verify launch crash resistance on low RAM emulation",
        "Device memory usage is artificially pushed near limit.",
        "1. Launch FormSahayak app.\n2. Monitor logs for memory warnings or crashes.",
        "RAM Limit: Low memory threshold",
        "App launches successfully and manages memory without throwing OutOfMemory errors.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"login_btn\")", "No",
        ["Regression"]
    ),
    (
        "App Installation", "Permissions Retention", "Verify camera permissions are retained after package update",
        "Older app version is installed. Camera permission is granted.",
        "1. Install upgrade APK over current install.\n2. Launch app and navigate to camera upload.\n3. Click camera button.",
        "N/A",
        "Camera opens immediately without re-prompting for permission.",
        "Medium", "Minor", "Appium Native", "id(\"com.android.permissioncontroller:id/permission_allow_button\")", "Yes",
        ["Regression"]
    ),
    (
        "App Installation", "Fresh Install", "Verify fresh installation via Google Play Store simulation",
        "FormSahayak is not installed.",
        "1. Open Play Store app on emulator.\n2. Search FormSahayak.\n3. Click Install.",
        "Simulated download",
        "App downloads and installs without errors.",
        "Medium", "Major", "Appium Native", "xpath(\"//*[@text='Install']\")", "No",
        ["Regression"]
    ),
    (
        "App Installation", "Upgrade Install", "Verify backward compatibility of local database on upgrade",
        "Older version installed with local database records (history).",
        "1. Install newer version.\n2. Launch app and open Form History tab.\n3. Check old records.",
        "Older DB records populated",
        "All historical data is preserved and readable. Migration script runs successfully.",
        "High", "Major", "Appium Native", "accessibilityId(\"history_item_0\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "App Installation", "App Launch", "Verify app launches correctly after system reboot",
        "App is installed.",
        "1. Reboot device.\n2. Tap FormSahayak icon.\n3. Check if app loads.",
        "N/A",
        "App launches cleanly post reboot.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"login_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "App Installation", "App Launch", "Verify launch crash resistance when background services fail",
        "Backend services simulated down.",
        "1. Block backend APIs at network level.\n2. Launch app.\n3. Verify app launches with offline mode alert.",
        "Backend offline",
        "App launches gracefully, showing no crash dialogs, just a network error overlay.",
        "High", "Major", "Appium Native", "accessibilityId(\"network_error_msg\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),

    # B. Registration Testing (20 TCs)
    (
        "Registration", "Valid Registration", "Verify registration with valid input details",
        "User is on the Registration screen.",
        "1. Enter unique username.\n2. Enter valid unused email.\n3. Enter valid password (strength rules met).\n4. Click 'Sign Up'.",
        "Username: 'vijay_k', Email: 'vijay.k@example.com', Password: 'SecuredPassword@123'",
        "User registered. Directed to Language Selection or Dashboard. Database writes record.",
        "High", "Critical", "Appium Native", "accessibilityId(\"register_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Registration", "Invalid Email", "Verify registration fails with invalid email format",
        "User is on the Registration screen.",
        "1. Enter valid username.\n2. Enter invalid email format.\n3. Enter valid password.\n4. Click 'Sign Up'.",
        "Email: 'vijay.k_example.com'",
        "Toast or error text under email field: 'Enter a valid email address.'",
        "High", "Major", "Appium Native", "xpath(\"//*[@text='Enter a valid email address.']\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Registration", "Duplicate Email", "Verify registration fails with an already registered email",
        "User is on the Registration screen. Email is already registered.",
        "1. Enter username.\n2. Enter existing email.\n3. Enter valid password.\n4. Click 'Sign Up'.",
        "Email: 'vijay.k@example.com'",
        "Error message: 'Email address already registered.'",
        "High", "Major", "Appium Native", "xpath(\"//*[@text='Email address already registered.']\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Registration", "Empty Fields", "Verify registration fails when all fields are empty",
        "User is on the Registration screen.",
        "1. Leave fields blank.\n2. Tap 'Sign Up'.",
        "N/A",
        "Validation messages: 'Fields cannot be empty.'",
        "High", "Major", "Appium Native", "xpath(\"//*[@text='Username is required']\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Registration", "Empty Username", "Verify registration fails with blank username field",
        "User is on the Registration screen.",
        "1. Leave username blank.\n2. Fill other details.\n3. Tap 'Sign Up'.",
        "Email: 'test@example.com', Password: 'Password@123'",
        "Validation error: 'Username is required.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[@text='Username is required.']\")", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "Empty Email", "Verify registration fails with blank email field",
        "User is on the Registration screen.",
        "1. Fill username.\n2. Leave email blank.\n3. Fill password.\n4. Tap 'Sign Up'.",
        "Username: 'vijay_k', Password: 'Password@123'",
        "Validation error: 'Email is required.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[@text='Email is required.']\")", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "Empty Password", "Verify registration fails with blank password field",
        "User is on the Registration screen.",
        "1. Fill username and email.\n2. Leave password blank.\n3. Tap 'Sign Up'.",
        "Username: 'vijay_k', Email: 'vijay.k@example.com'",
        "Validation error: 'Password is required.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[@text='Password is required.']\")", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "Password Length", "Verify registration fails with password under 8 characters",
        "User is on the Registration screen.",
        "1. Enter username and email.\n2. Enter short password (6 chars).\n3. Click 'Sign Up'.",
        "Password: 'Pass12'",
        "Validation error: 'Password must be at least 8 characters long.'",
        "High", "Major", "Appium Native", "xpath(\"//*[@text='Password must be at least 8 characters long.']\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Registration", "Password Casing", "Verify registration fails with password lacking uppercase letters",
        "User is on the Registration screen.",
        "1. Enter details.\n2. Enter password lacking uppercase.\n3. Click 'Sign Up'.",
        "Password: 'securedpassword@123'",
        "Validation error: 'Password must contain at least one uppercase letter.'",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'uppercase')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "Password Numbers", "Verify registration fails with password lacking numeric digits",
        "User is on the Registration screen.",
        "1. Enter details.\n2. Enter password with letters only.\n3. Click 'Sign Up'.",
        "Password: 'SecuredPassword@'",
        "Validation error: 'Password must contain at least one digit.'",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'digit')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "Password Special Characters", "Verify registration fails with password lacking special characters",
        "User is on the Registration screen.",
        "1. Enter details.\n2. Enter password with letters and numbers only.\n3. Click 'Sign Up'.",
        "Password: 'SecuredPassword123'",
        "Validation error: 'Password must contain at least one special character.'",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'special character')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "Password Match", "Verify registration fails if confirm password mismatch",
        "User is on the Registration screen.",
        "1. Enter password.\n2. Enter mismatching confirm password.\n3. Click 'Sign Up'.",
        "Password: 'Password@123', Confirm: 'Password@124'",
        "Validation error: 'Passwords do not match.'",
        "High", "Major", "Appium Native", "xpath(\"//*[@text='Passwords do not match.']\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Registration", "Input Sanitization", "Verify SQL injection check in Registration fields",
        "User is on the Registration screen.",
        "1. Enter SQL injection string in username.\n2. Tap 'Sign Up'.",
        "Username: \"admin' OR '1'='1\"",
        "Input is sanitized. No database exception; registration either proceeds as literal username string or errors gracefully.",
        "High", "Critical", "Appium Native", "accessibilityId(\"register_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Registration", "Input Sanitization", "Verify XSS scripts are stripped from Username field",
        "User is on the Registration screen.",
        "1. Enter JS alert script tag in username.\n2. Complete registration.\n3. View registered username on profile dashboard.",
        "Username: \"<script>alert('xss')</script>\"",
        "Tags are HTML encoded or stripped. No script execution occurs in Compose views.",
        "High", "Critical", "Appium Native", "accessibilityId(\"profile_name_text\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Registration", "Spaces Handling", "Verify email spaces are trimmed before submission",
        "User is on the Registration screen.",
        "1. Enter email with leading and trailing spaces.\n2. Fill valid details and tap 'Sign Up'.",
        "Email: '  vijay.k@example.com  '",
        "Registration succeeds. Checked DB shows email stored without spaces.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"register_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "Keyboard Behavior", "Verify done button on soft keyboard triggers submission",
        "User is on the password field during registration.",
        "1. Focus on Confirm Password field.\n2. Tap 'Done' / 'Enter' on soft keyboard.\n3. Verify form submissions.",
        "Soft keyboard active",
        "Keyboard hides and signup API is triggered.",
        "Medium", "Minor", "Appium Native", "N/A (Keyboard action)", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "Terms Link", "Verify redirection to Privacy Policy / Terms of Service link",
        "User is on the Registration screen.",
        "1. Click the 'Terms of Service' text link.\n2. Check redirection.",
        "N/A",
        "Browser opens containing terms of service text, app goes to background.",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'Terms')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "Password Masking", "Verify toggle password visibility works",
        "User is typing password.",
        "1. Enter password.\n2. Verify text is masked (dots).\n3. Click eye icon.\n4. Verify text is visible.",
        "N/A",
        "Password text transitions from bullets to clear text and back to bullets.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"password_visibility_toggle\")", "Yes",
        ["Sanity", "Regression"]
    ),
    (
        "Registration", "Valid Registration", "Verify registration preserves case sensitivity for username",
        "User registering with mixed casing.",
        "1. Register username 'VijayK'.\n2. Open profile screen after success.",
        "Username: 'VijayK'",
        "Username displayed exactly as 'VijayK' on screen.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"profile_name_text\")", "Yes",
        ["Regression"]
    ),
    (
        "Registration", "API Connection Fail", "Verify error handling when backend is offline during registration",
        "User clicks registration button while backend is simulated offline.",
        "1. Fill all registration details.\n2. Click 'Sign Up' when server is offline.\n3. Check error output.",
        "N/A",
        "Error message displayed: 'Unable to connect to server. Please try again later.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'connect to server')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),

    # C. Login Testing (20 TCs)
    (
        "Login", "Valid Login", "Verify successful login with valid registered email and password",
        "User is registered. Active on the login screen.",
        "1. Enter correct email.\n2. Enter correct password.\n3. Click 'Login'.",
        "Email: 'vijay.k@example.com', Password: 'SecuredPassword@123'",
        "Dashboard screen is loaded. User session is saved in local shared preferences.",
        "High", "Critical", "Appium Native", "accessibilityId(\"login_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Login", "Invalid Credentials", "Verify login fails with unregistered email",
        "User is on the login screen.",
        "1. Enter unregistered email.\n2. Enter password.\n3. Tap 'Login'.",
        "Email: 'notexist@example.com', Password: 'SecuredPassword@123'",
        "Error message: 'Invalid email or password.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Invalid')]\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Login", "Invalid Credentials", "Verify login fails with correct email but incorrect password",
        "User is registered. Active on the login screen.",
        "1. Enter registered email.\n2. Enter incorrect password.\n3. Tap 'Login'.",
        "Email: 'vijay.k@example.com', Password: 'WrongPassword'",
        "Error message: 'Invalid email or password.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Invalid')]\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Login", "Empty Fields", "Verify login fails when fields are empty",
        "User is on the login screen.",
        "1. Leave fields blank.\n2. Tap 'Login'.",
        "N/A",
        "Validation errors: 'Email and Password are required.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'required')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Login", "Session Persistence", "Verify session remains active after closing and reopening the app",
        "User is logged in on the dashboard screen.",
        "1. Close the application (kill process).\n2. Launch application again.\n3. Verify dashboard screen loads directly.",
        "N/A",
        "Bypasses login screen. User is taken straight to the dashboard.",
        "High", "Critical", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Login", "Logout Validation", "Verify successful logout from application",
        "User is logged in and is on the dashboard screen.",
        "1. Navigate to Profile settings.\n2. Click the 'Logout' button.\n3. Confirm logout.",
        "N/A",
        "Session invalidated, local preferences cleared, user redirected to Login screen.",
        "High", "Critical", "Appium Native", "accessibilityId(\"logout_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Login", "Logout Validation", "Verify back button cannot access dashboard after logout",
        "User has logged out and is on the login screen.",
        "1. Press device hardware back button.",
        "N/A",
        "App remains on the login screen or exits. Does not return to the dashboard.",
        "High", "Critical", "Appium Native", "accessibilityId(\"login_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Login", "Input Sanitization", "Verify SQL injection prevention in login inputs",
        "User is on the login screen.",
        "1. Input SQL injection query in email.\n2. Input string in password.\n3. Tap 'Login'.",
        "Email: \"' OR 1=1 --\"",
        "Server rejects query safely; handles input as a standard string. Showing invalid credentials.",
        "High", "Critical", "Appium Native", "accessibilityId(\"login_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Login", "Forgot Password", "Verify redirection and flow of Forgot Password link",
        "User is on the login screen.",
        "1. Tap 'Forgot Password' link.\n2. Enter registered email.\n3. Tap 'Send Reset Link'.",
        "Email: 'vijay.k@example.com'",
        "App displays success alert: 'Password reset link sent to your email.'",
        "High", "Major", "Appium Native", "accessibilityId(\"forgot_password_btn\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Login", "Forgot Password", "Verify Forgot Password fails with unregistered email",
        "User is on Forgot Password screen.",
        "1. Enter unregistered email.\n2. Tap 'Send Reset Link'.",
        "Email: 'notregistered@example.com'",
        "Error alert displayed: 'Email address not found.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'not found')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Login", "Account Lockout", "Verify account lockout after multiple failed login attempts",
        "User is registered.",
        "1. Input incorrect password 5 consecutive times.\n2. Observe error status and message.",
        "Email: 'vijay.k@example.com', Password: 'WrongPassword'",
        "Account is locked temporarily for 15 minutes. Error displays: 'Too many attempts. Try again in 15 minutes.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Too many attempts')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Login", "Case Insensitivity", "Verify login is case-insensitive for email field",
        "User registered email is lowercase.",
        "1. Input email in uppercase characters.\n2. Enter correct password.\n3. Tap 'Login'.",
        "Email: 'VIJAY.K@EXAMPLE.COM', Password: 'SecuredPassword@123'",
        "Dashboard screen is loaded. User logs in successfully.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"login_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Login", "Case Sensitivity", "Verify login is case-sensitive for password field",
        "User is registered.",
        "1. Input registered email.\n2. Input password with incorrect casing.\n3. Tap 'Login'.",
        "Email: 'vijay.k@example.com', Password: 'securedpassword@123'",
        "Error message: 'Invalid email or password.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Invalid')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Login", "Session Persistence", "Verify session is invalidated when user logs in from another device",
        "User is logged in on Device A.",
        "1. Log in with same credentials on Device B.\n2. Perform activity on Device A.\n3. Verify Device A behavior.",
        "N/A",
        "Device A session terminates. Automatically redirects to login with message 'Logged in from another device.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Logged in')]\")", "No",
        ["Regression"]
    ),
    (
        "Login", "Session Persistence", "Verify session is preserved if device loses connection briefly",
        "User is logged in on the dashboard screen.",
        "1. Toggle internet connection off for 10 seconds.\n2. Toggle internet back on.\n3. Perform transaction.",
        "N/A",
        "User is not logged out. Session remains active.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Regression"]
    ),
    (
        "Login", "Remember Me", "Verify 'Remember Me' retains email input box value",
        "User is on the login screen.",
        "1. Enter email.\n2. Check 'Remember Me' checkbox.\n3. Log in successfully.\n4. Log out.\n5. Verify login screen.",
        "Email: 'vijay.k@example.com'",
        "Email text field is auto-populated with 'vijay.k@example.com'.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"remember_me_checkbox\")", "Yes",
        ["Regression"]
    ),
    (
        "Login", "Token Storage", "Verify token is stored securely in encrypted shared preferences",
        "User has logged in successfully.",
        "1. Check app private folders inside emulator using root command.\n2. Inspect shared preference XML content.",
        "N/A",
        "Authentication tokens are encrypted and cannot be read as plain text.",
        "High", "Critical", "Appium Native", "N/A (Security audit command)", "No",
        ["Security", "Regression"]
    ),
    (
        "Login", "Auto-Focus", "Verify cursor is auto-focused on email field upon opening login screen",
        "User opens login screen.",
        "1. Open the login screen.\n2. Observe keyboard status and cursor.",
        "N/A",
        "Cursor is blinking in the email field, keyboard is displayed.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"login_email\")", "Yes",
        ["UI/UX", "Regression"]
    ),
    (
        "Login", "Password Masking", "Verify toggle password visibility works on login screen",
        "User enters password.",
        "1. Type password.\n2. Tap eye toggle button.\n3. Tap eye toggle button again.",
        "N/A",
        "Password toggles between bullets and plain text.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"login_password_visibility\")", "Yes",
        ["Sanity", "Regression"]
    ),
    (
        "Login", "Offline Login Fail", "Verify login fails with warning when device is completely offline",
        "User is on the login screen. Device has no internet.",
        "1. Disable cellular data and Wi-Fi.\n2. Enter credentials.\n3. Tap 'Login'.",
        "N/A",
        "Toast message displayed: 'Offline. Please check your internet connection.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'check your internet')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),

    # D. Profile Testing (20 TCs)
    (
        "Profile", "Update Profile", "Verify updating username with valid characters",
        "User is logged in. Navigates to Edit Profile.",
        "1. Click 'Edit Profile'.\n2. Change name.\n3. Tap 'Save'.",
        "New Name: 'Vijay Kumar'",
        "Name updated in database and reflects immediately on the profile dashboard.",
        "High", "Major", "Appium Native", "accessibilityId(\"save_profile_btn\")", "Yes",
        ["Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Profile", "Update Profile", "Verify profile editing fails if name field is cleared",
        "User is on Edit Profile screen.",
        "1. Clear name field.\n2. Tap 'Save'.",
        "N/A",
        "Error validation: 'Name cannot be blank.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'cannot be blank')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Profile", "Change Language", "Verify app language changes to English",
        "User is on profile language settings screen.",
        "1. Tap language dropdown.\n2. Select 'English'.",
        "Language: English",
        "App UI immediately re-renders all static text to English.",
        "High", "Critical", "Appium Native", "xpath(\"//*[@text='English']\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Profile", "Change Language", "Verify app language changes to Telugu",
        "User is on profile language settings screen.",
        "1. Tap language dropdown.\n2. Select 'Telugu' (తెలుగు).",
        "Language: Telugu",
        "App UI immediately re-renders all static text to Telugu translation.",
        "High", "Critical", "Appium Native", "xpath(\"//*[@text='తెలుగు']\")", "Yes",
        ["Smoke", "Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Profile", "Change Language", "Verify app language changes to Hindi",
        "User is on profile language settings screen.",
        "1. Tap language dropdown.\n2. Select 'Hindi' (हिन्दी).",
        "Language: Hindi",
        "App UI immediately re-renders all static text to Hindi translation.",
        "High", "Critical", "Appium Native", "xpath(\"//*[@text='हिन्दी']\")", "Yes",
        ["Smoke", "Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Profile", "Change Language", "Verify app language changes to Tamil",
        "User is on profile language settings screen.",
        "1. Tap language dropdown.\n2. Select 'Tamil' (தமிழ்).",
        "Language: Tamil",
        "App UI immediately re-renders all static text to Tamil translation.",
        "High", "Critical", "Appium Native", "xpath(\"//*[@text='தமிழ்']\")", "Yes",
        ["Smoke", "Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify uploading profile image in valid JPG format",
        "User is logged in. Navigates to Edit Profile.",
        "1. Tap profile avatar.\n2. Select a valid JPG image from gallery.\n3. Tap 'Save'.",
        "File: profile.jpg (1MB)",
        "Image is uploaded to cloud database, avatar reflects new image.",
        "High", "Major", "Appium Native", "accessibilityId(\"profile_avatar\")", "Yes",
        ["Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify uploading profile image in PNG format",
        "User is logged in.",
        "1. Tap avatar.\n2. Select valid PNG.\n3. Tap 'Save'.",
        "File: avatar.png",
        "Image is uploaded and rendered successfully.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"profile_avatar\")", "Yes",
        ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify image size exceeds limit warning",
        "User is logged in. Limit is 5MB.",
        "1. Select image of 6.2MB size.\n2. Attempt upload.",
        "File: heavy_image.jpg (6.2MB)",
        "Error message displayed: 'File size exceeds maximum limit of 5MB.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'exceeds')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify upload fails with unsupported extension format (GIF)",
        "User is logged in.",
        "1. Select a GIF format file from storage.\n2. Attempt upload.",
        "File: animated.gif",
        "Error toast displayed: 'Invalid format. Only JPG and PNG allowed.'",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'Only JPG and PNG')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Profile", "Remove Profile Image", "Verify profile image removal resets to placeholder avatar",
        "User has a custom profile image uploaded.",
        "1. Tap edit avatar.\n2. Select 'Remove Image'.\n3. Tap 'Save'.",
        "N/A",
        "Image is deleted from storage server. Avatar resets to default initials placeholder.",
        "High", "Major", "Appium Native", "accessibilityId(\"remove_avatar_btn\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Profile", "Profile Persistence", "Verify changes persist across logging out and logging back in",
        "User is logged in.",
        "1. Update name to 'Vijay K'.\n2. Log out.\n3. Log in with same account.\n4. View Profile dashboard.",
        "N/A",
        "Name displays as 'Vijay K' on Profile screen.",
        "High", "Major", "Appium Native", "accessibilityId(\"profile_name_text\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Profile", "Update Phone", "Verify updating phone number with valid 10-digit number",
        "User is on Edit Profile screen.",
        "1. Enter valid 10 digit number.\n2. Tap 'Save'.",
        "Phone: '9876543210'",
        "Phone number is successfully saved and displayed in profile details.",
        "High", "Major", "Appium Native", "accessibilityId(\"phone_input\")", "Yes",
        ["Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Profile", "Update Phone", "Verify phone update validation error with short number",
        "User is on Edit Profile screen.",
        "1. Enter 8 digit phone number.\n2. Tap 'Save'.",
        "Phone: '98765432'",
        "Validation error: 'Phone number must be 10 digits.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, '10 digits')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Profile", "Update Phone", "Verify phone update validation error with alphabetic characters",
        "User is on Edit Profile screen.",
        "1. Enter phone number containing letters.\n2. Tap 'Save'.",
        "Phone: '98765432ab'",
        "Input is either blocked or shows warning: 'Invalid phone number.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Invalid')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify upload fails with empty file (0 bytes)",
        "User is on edit profile.",
        "1. Attempt to upload a mock empty file of 0 bytes.\n2. Check app behavior.",
        "File size: 0 bytes",
        "Error message displayed: 'Uploaded file is empty.'",
        "Low", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'empty')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Profile", "Upload Profile Image", "Verify upload behavior when server returns timeout (Coil loading)",
        "Slow network simulation active.",
        "1. Select profile picture.\n2. Tap Upload.\n3. Simulate server timeout.",
        "Timeout: 30 seconds",
        "App shows loader and then handles timeout gracefully: 'Connection timed out. Try again.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'timed out')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Profile", "Cancel Changes", "Verify clicking Cancel button discards all profile changes",
        "User modifies details on Edit Profile screen.",
        "1. Change name field content.\n2. Tap 'Cancel' or back button.\n3. Return to view profile.",
        "Name changed to 'Temp Name'",
        "Profile name remains unchanged (keeps previous value).",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"cancel_profile_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Profile", "Security Audit", "Verify password fields are masked during profile update",
        "User is modifying credentials.",
        "1. Enter password fields.\n2. Check layout.",
        "N/A",
        "Passwords remain masked by default.",
        "High", "Major", "Appium Native", "accessibilityId(\"password_input\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Profile", "Input Sanitization", "Verify profile update input sanitization for name",
        "User is on Edit Profile screen.",
        "1. Input raw HTML tag elements in name field.\n2. Tap 'Save'.",
        "Name: \"<b>Vijay</b>\"",
        "Name stores as plain text or tags are stripped. UI renders name literally as '<b>Vijay</b>' or 'Vijay'. No HTML formatting parsed.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"profile_name_text\")", "Yes",
        ["Regression"]
    ),

    # E. Camera & Permissions Testing (15 TCs)
    (
        "Camera & Permissions", "Camera Permission Allow", "Verify camera opens successfully when permission is allowed",
        "Camera permission is not yet prompt state.",
        "1. Tap 'Upload Form' -> 'Take Photo'.\n2. Dialog displays. Tap 'Allow' / 'While using the app'.\n3. Observe camera viewer.",
        "Permission: Allow",
        "Camera view opens, permitting user to snap picture of form.",
        "High", "Critical", "Appium Native", "xpath(\"//*[@text='Allow' or contains(@text, 'While using')]\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Camera & Permissions", "Camera Permission Deny", "Verify error message when camera permission is denied",
        "Camera permission is not yet prompt state.",
        "1. Tap 'Upload Form' -> 'Take Photo'.\n2. Dialog displays. Tap 'Deny' / 'Don't Allow'.\n3. Observe behavior.",
        "Permission: Deny",
        "Camera view does not open. Toast message displays: 'Camera permission is required to capture forms.'",
        "High", "Critical", "Appium Native", "xpath(\"//*[@text='Deny' or contains(@text, 'Don')]\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Camera & Permissions", "Storage Permission Allow", "Verify gallery opens successfully when storage permission is allowed",
        "Storage permission is in unprompted state.",
        "1. Tap 'Upload Form' -> 'Choose from Gallery'.\n2. Tap 'Allow' / 'Allow Access to Media'.\n3. Observe image list.",
        "Permission: Allow",
        "Device file picker/gallery lists images successfully.",
        "High", "Critical", "Appium Native", "xpath(\"//*[@text='Allow']\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Camera & Permissions", "Storage Permission Deny", "Verify warning when storage permission is denied",
        "Storage permission is in unprompted state.",
        "1. Tap 'Upload Form' -> 'Choose from Gallery'.\n2. Tap 'Deny'.\n3. Observe app layout.",
        "Permission: Deny",
        "Gallery does not open. Toast warning: 'Storage permission is required to upload forms.'",
        "High", "Critical", "Appium Native", "xpath(\"//*[@text='Deny']\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Camera & Permissions", "Microphone Permission Allow", "Verify microphone icon is enabled when microphone permission is allowed",
        "Microphone permission not prompted.",
        "1. Tap 'Voice Guidance' micro icon.\n2. Tap 'Allow'.\n3. Verify microphone activity indicator.",
        "Permission: Allow",
        "Microphone active state; user can speak commands or query guidance voice.",
        "High", "Major", "Appium Native", "xpath(\"//*[@text='Allow']\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Camera & Permissions", "Microphone Permission Deny", "Verify alert message when microphone permission is denied",
        "Microphone permission not prompted.",
        "1. Tap 'Voice Guidance' mic button.\n2. Tap 'Deny'.\n3. Verify voice interface output.",
        "Permission: Deny",
        "Voice input disabled. Message appears: 'Microphone permission required for voice guidance inputs.'",
        "High", "Major", "Appium Native", "xpath(\"//*[@text='Deny']\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Camera & Permissions", "Permission Revoked", "Verify app behavior when camera permission is revoked mid-session via OS Settings",
        "App is running in background. Camera permission was allowed.",
        "1. Open device system settings.\n2. Revoke Camera permission for FormSahayak.\n3. Return to app.\n4. Click 'Take Photo'.",
        "Revoke permission via OS Settings",
        "App prompts again for permission or handles background process restart gracefully without crash.",
        "High", "Major", "Appium Native", "xpath(\"//*[@text='Allow' or contains(@text, 'While using')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Camera & Permissions", "Permission Revoked", "Verify app behavior when storage permission is revoked mid-session",
        "App running in background. Storage permission allowed.",
        "1. Revoke Storage permission from OS settings.\n2. Resume app.\n3. Tap 'Choose from Gallery'.",
        "Revoke permission",
        "App prompts user again for storage permission.",
        "Medium", "Major", "Appium Native", "xpath(\"//*[@text='Allow']\")", "Yes",
        ["Regression"]
    ),
    (
        "Camera & Permissions", "Camera Permission Deny", "Verify 'Don't ask again' camera deny directs to Settings",
        "Camera permission is prompted.",
        "1. Tap 'Deny and don't ask again'.\n2. Click 'Take Photo' again.\n3. Observe dialog prompt.",
        "Permission: Deny permanent",
        "A dialog appears showing: 'Camera permission permanently denied. Enable it in Settings.' with a direct link button.",
        "High", "Major", "Appium Native", "accessibilityId(\"go_to_settings_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Camera & Permissions", "Storage Permission Deny", "Verify 'Don't ask again' storage deny directs to Settings",
        "Storage permission prompted.",
        "1. Tap 'Deny and don't ask again'.\n2. Click 'Choose from Gallery' again.\n3. Observe dialog.",
        "Permission: Deny permanent",
        "Dialog directs user to Settings with link button.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"go_to_settings_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Camera & Permissions", "Camera Permission Allow", "Verify camera functionality with front and back camera toggles",
        "Camera open state.",
        "1. Click switch camera icon.\n2. Check camera orientation.",
        "Camera view active",
        "App switches between back and front camera feed smoothly.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"switch_camera_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Camera & Permissions", "Camera Permission Allow", "Verify flash toggle button behavior",
        "Camera open state.",
        "1. Tap flash toggle button.\n2. Capture picture.",
        "Flash Mode: ON",
        "Device flash fires during photo capture.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"flash_toggle_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Camera & Permissions", "Permission Flow", "Verify system dialog trigger flow inside Compose wizard",
        "Clean permissions state.",
        "1. Initiate document capture path.\n2. Verify system dialog overlays Compose screen.",
        "N/A",
        "System permission dialog appears seamlessly over Compose UI.",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'Allow')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Camera & Permissions", "Microphone Permission Allow", "Verify microphone permission state remains on phone lock",
        "Microphone permission allowed.",
        "1. Lock screen.\n2. Unlock screen.\n3. Test voice command.",
        "N/A",
        "Microphone permission remains active, no new prompt required.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"voice_mic_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Camera & Permissions", "Storage Permission Allow", "Verify picking file from external SD card storage",
        "Storage permission allowed.",
        "1. Navigate file picker.\n2. Select file from SD card.",
        "File: sdcard/form.jpg",
        "File selected and uploaded successfully.",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'SD Card') or contains(@text, 'External')]\")", "No",
        ["Regression"]
    ),

    # F. OCR Testing (20 TCs)
    (
        "OCR", "Clear Image OCR", "Verify EasyOCR text extraction on clear high-resolution English document",
        "User is logged in. Document uploaded is a clear English printed form.",
        "1. Tap 'Extract Text' / 'Run OCR'.\n2. Wait for API to respond.\n3. Check detected text fields.",
        "Form: sbi_opening_form.jpg (300 DPI)",
        "EasyOCR parses characters accurately (above 95% correctness). Text is mapped to correct coordinate grids.",
        "High", "Critical", "Appium Native", "accessibilityId(\"run_ocr_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "OCR", "Clear Image OCR", "Verify OCR text extraction on clean Telugu document",
        "User uploaded clean Telugu document.",
        "1. Tap 'Run OCR'.\n2. Inspect extracted overlay.",
        "Form: telugu_form.png",
        "OCR returns Telugu text correctly, showing translated text in the guided fields.",
        "High", "Critical", "Appium Native", "accessibilityId(\"run_ocr_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "OCR", "Clear Image OCR", "Verify OCR text extraction on clean Hindi document",
        "User uploaded clean Hindi document.",
        "1. Tap 'Run OCR'.\n2. Verify output.",
        "Form: hindi_form.jpg",
        "Hindi characters parsed and extracted correctly.",
        "High", "Critical", "Appium Native", "accessibilityId(\"run_ocr_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "OCR", "Clear Image OCR", "Verify OCR text extraction on clean Tamil document",
        "User uploaded clean Tamil document.",
        "1. Tap 'Run OCR'.\n2. Verify output.",
        "Form: tamil_form.jpg",
        "Tamil characters parsed and extracted correctly.",
        "High", "Critical", "Appium Native", "accessibilityId(\"run_ocr_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "OCR", "Blurry Image OCR", "Verify OCR display warning on low quality blurry image",
        "Uploaded image has high blur index.",
        "1. Click 'Run OCR'.\n2. Observe toast or error overlay.",
        "File: blurry_pic.jpg",
        "Text extracted if possible, but warning appears: 'Image quality is low. Field suggestions may be inaccurate.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'quality is low')]\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "OCR", "Blurry Image OCR", "Verify OCR fails completely on unreadable extremely blurry image",
        "Uploaded image is totally pixelated/blurry.",
        "1. Click 'Run OCR'.\n2. Verify behavior.",
        "File: bad_blur.jpg",
        "Error message displayed: 'OCR failed. Unable to read document. Please upload a clearer image.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Unable to read document')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "OCR", "Rotated Image OCR", "Verify OCR on 90-degree rotated document auto-corrects orientation",
        "Uploaded image is rotated 90 degrees.",
        "1. Tap 'Run OCR'.\n2. Inspect extracted bounding boxes.",
        "File: rotated_90.png",
        "Server rotates image automatically or reads text in correct vertical alignment.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"ocr_overlay\")", "Yes",
        ["Regression"]
    ),
    (
        "OCR", "Rotated Image OCR", "Verify OCR on 180-degree upside down document corrects orientation",
        "Uploaded image is rotated 180 degrees.",
        "1. Tap 'Run OCR'.\n2. Verify output text layout.",
        "File: inverted_180.png",
        "Server processes upright layout and highlights boxes at original positions.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"ocr_overlay\")", "Yes",
        ["Regression"]
    ),
    (
        "OCR", "Large Image OCR", "Verify OCR processing stability on massive dimensions image file",
        "Image size is 4.8MB (near limit).",
        "1. Upload 4.8MB JPG.\n2. Tap 'Run OCR'.\n3. Wait for extraction.",
        "File: huge_dimensions.jpg",
        "OCR succeeds without server memory crashes. Returns bounds in under 10 seconds.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"run_ocr_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "OCR", "Empty Image OCR", "Verify OCR fails with proper alert on a completely blank page",
        "Uploaded image is completely white or blank.",
        "1. Click 'Run OCR'.\n2. Observe alert dialog.",
        "File: blank_canvas.jpg",
        "Error displayed: 'No structured text detected. Please upload a valid document form.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'No structured text')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "OCR", "Bilingual OCR", "Verify OCR on mixed language (English and Telugu) form",
        "Uploaded image contains both English and Telugu text.",
        "1. Tap 'Run OCR'.\n2. Verify coordinates and language fonts.",
        "File: bilingual_card.jpg",
        "EasyOCR parses both languages and displays bilingual tooltips.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"ocr_overlay\")", "Yes",
        ["Regression"]
    ),
    (
        "OCR", "Cancel Operation", "Verify user can abort OCR execution mid-process",
        "OCR is in progress state.",
        "1. Tap 'Cancel' on the loading overlay.\n2. Verify loader ceases.",
        "N/A",
        "API call is aborted, spinner disappears, UI returns to upload view.",
        "High", "Major", "Appium Native", "accessibilityId(\"cancel_loading_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "OCR", "OCR Server Latency", "Verify retry toast display on server-side timeout",
        "Simulating backend delay of 45 seconds.",
        "1. Click 'Run OCR'.\n2. Wait for client timeout limit (30s).",
        "Server timeout simulation",
        "Loader stops, error message shows: 'Request timed out. Please try again.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'timed out')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "OCR", "Non-Document OCR", "Verify app behavior when uploading photo of landscape/object",
        "User uploads image of a desk or landscape.",
        "1. Click 'Run OCR'.\n2. Observe classification output.",
        "File: landscape.jpg",
        "App identifies non-document image and errors: 'Uploaded file is not identified as a form.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'not identified as a form')]\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "OCR", "OCR Coordinates", "Verify bounding box layout aligns with actual image text fields",
        "OCR returns coordinate JSON.",
        "1. Observe Compose highlight overlays on image viewport.",
        "N/A",
        "Coordinates overlap the actual printed text exactly, without offset displacements.",
        "High", "Major", "Appium Native", "accessibilityId(\"highlight_box_1\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "OCR", "Handwritten Text", "Verify OCR filters out handwritten fields from structure mapping",
        "Uploaded form contains print structures filled with pen.",
        "1. Click 'Run OCR'.\n2. Observe highlighted labels.",
        "File: filled_form.jpg",
        "Print labels are highlighted for guidance; handwritten text does not confuse structure mapper.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"ocr_overlay\")", "Yes",
        ["Regression"]
    ),
    (
        "OCR", "Pinch-To-Zoom OCR", "Verify coordinate overlays scale correctly during pinch-to-zoom",
        "OCR overlays active on screen.",
        "1. Perform pinch-to-zoom on the form image.\n2. Verify highlight boxes positions.",
        "Pinch gesture",
        "Highlight boxes scale dynamically and remain centered on their relative text elements.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"ocr_overlay\")", "Yes",
        ["Regression"]
    ),
    (
        "OCR", "Shadow Image OCR", "Verify OCR readability under heavy uneven shadow",
        "Form photo taken under harsh shadow.",
        "1. Click 'Run OCR'.\n2. Verify extraction success rate.",
        "File: shadow_form.jpg",
        "Pre-processing handles thresholding, text is extracted with > 80% accuracy.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"ocr_overlay\")", "No",
        ["Regression"]
    ),
    (
        "OCR", "Folded Image OCR", "Verify OCR readability on wrinkled or folded paper form",
        "Form uploaded is heavily folded.",
        "1. Click 'Run OCR'.\n2. Inspect labels.",
        "File: folded.png",
        "Form structure recognized and key guides generated.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"ocr_overlay\")", "No",
        ["Regression"]
    ),
    (
        "OCR", "OCR JSON Parsing", "Verify app handles corrupted JSON response from OCR backend",
        "Backend sends malformed JSON simulation.",
        "1. Click 'Run OCR'.\n2. Inject corrupt response payload.\n3. Check client status.",
        "Simulated response: '{invalid_json}'",
        "App does not crash; displays: 'Failed to process server response. Please try again.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Failed to process')]\")", "Yes",
        ["Regression"]
    ),

    # G. Form Upload Testing (20 TCs)
    (
        "Form Upload", "JPG Upload", "Verify successful upload of JPG form document",
        "User is logged in. Navigates to Upload Form page.",
        "1. Select a JPG form.\n2. Tap 'Upload'.\n3. Wait for success response.",
        "File: form.jpg (1.8MB)",
        "Form upload API returns 201. File is saved. Redirected to OCR step.",
        "High", "Critical", "Appium Native", "accessibilityId(\"upload_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Form Upload", "PNG Upload", "Verify successful upload of PNG form document",
        "User is on Upload page.",
        "1. Select a PNG form.\n2. Tap 'Upload'.",
        "File: form.png (900KB)",
        "Upload succeeds. Status code 201 returned.",
        "High", "Critical", "Appium Native", "accessibilityId(\"upload_btn\")", "Yes",
        ["Smoke", "Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "PDF Upload", "Verify successful upload of single-page PDF form",
        "User is on Upload page.",
        "1. Select a PDF file.\n2. Tap 'Upload'.",
        "File: application.pdf (single page)",
        "Upload succeeds. Backend extracts page as raster image for OCR processing.",
        "High", "Critical", "Appium Native", "accessibilityId(\"upload_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Form Upload", "PDF Upload", "Verify upload of multi-page PDF document",
        "User is on Upload page.",
        "1. Select a 3-page PDF form.\n2. Tap 'Upload'.",
        "File: passport.pdf (3 pages)",
        "Upload succeeds. Backend lists separate pages, enabling user to choose page for guidance.",
        "High", "Major", "Appium Native", "accessibilityId(\"upload_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "Max Size Limit", "Verify upload fails when file exceeds 5MB size limit",
        "User is on Upload page.",
        "1. Select a 5.8MB PNG file.\n2. Click 'Upload'.",
        "File: heavy_form.png (5.8MB)",
        "App blocks upload immediately. Toast: 'File exceeds maximum limit of 5MB.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'exceeds')]\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "Empty File", "Verify upload fails with 0-byte file",
        "User is on Upload page.",
        "1. Select a mock 0-byte file.\n2. Click 'Upload'.",
        "File: empty.jpg (0 bytes)",
        "App blocks upload. Message: 'Invalid file. File is empty.'",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'empty')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Form Upload", "Corrupted File", "Verify upload fails with corrupted PDF file",
        "User is on Upload page.",
        "1. Select a corrupted PDF file.\n2. Click 'Upload'.",
        "File: corrupt.pdf",
        "Upload fails. Message: 'Failed to upload. File is corrupted.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'corrupted')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Form Upload", "Invalid Format", "Verify upload fails with unsupported extension (.txt)",
        "User is on Upload page.",
        "1. Choose a txt file from file picker.\n2. Attempt upload.",
        "File: notes.txt",
        "File picker filters out or app displays: 'Unsupported file format. Only JPG, PNG, and PDF allowed.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Unsupported')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "Invalid Format", "Verify upload fails with unsupported extension (.zip)",
        "User is on Upload page.",
        "1. Try to upload a ZIP file.\n2. Observe response.",
        "File: package.zip",
        "Upload blocked. Shows error message.",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Unsupported')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "Invalid Format", "Verify upload fails with executable script (.apk / .sh)",
        "User attempts malicious file bypass.",
        "1. Select file 'malicious.sh' renamed to 'malicious.jpg' but holding shell commands.\n2. Tap 'Upload'.",
        "File: malicious.jpg (shell header)",
        "Server validates MIME type (magic numbers). Upload rejected: 'Invalid image format.'",
        "High", "Critical", "Appium Native", "xpath(\"//*[contains(@text, 'Invalid')]\")", "Yes",
        ["Security", "Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "Double Extension", "Verify file name with double extension (.pdf.jpg)",
        "User is on Upload page.",
        "1. Rename file to 'test.pdf.jpg'.\n2. Click 'Upload'.",
        "File: test.pdf.jpg",
        "Upload succeeds if binary structure is a clean JPG image.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"upload_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Form Upload", "Special Characters", "Verify file name with special characters upload successfully",
        "User is on Upload page.",
        "1. Select file 'bank_form_#$123!.png'.\n2. Click 'Upload'.",
        "File: bank_form_#$123!.png",
        "Upload succeeds. Filename is sanitized on server side before storage.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"upload_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Form Upload", "Progress Indicator", "Verify dynamic progress bar updates during transfer",
        "Uploading a large 4.5MB file.",
        "1. Tap 'Upload'.\n2. Monitor progress bar overlay.",
        "File: heavy_form.jpg (4.5MB)",
        "Progress bar fills smoothly from 0% to 100% correlating with network packets sent.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"upload_progress_bar\")", "Yes",
        ["UI/UX", "Regression"]
    ),
    (
        "Form Upload", "Cancel Upload", "Verify cancel button stops active upload",
        "User starts upload.",
        "1. Tap upload.\n2. Tap cancel/abort button mid-transfer.\n3. Verify UI.",
        "Upload in progress",
        "Upload terminates immediately. Incomplete file purged from server.",
        "High", "Major", "Appium Native", "accessibilityId(\"cancel_upload_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "Offline Block", "Verify upload is blocked immediately when offline",
        "User is on Upload page. Device network is disconnected.",
        "1. Select form.\n2. Tap 'Upload'.",
        "Network State: Offline",
        "App blocks request without calling API. Alerts: 'No network. Unable to upload.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'No network')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "Network Timeout", "Verify upload failure handling on API timeout",
        "Simulating network connection drop mid-way.",
        "1. Tap Upload.\n2. Terminate connection immediately.\n3. Observe behavior.",
        "N/A",
        "App handles upload timeout, offering 'Retry Upload' button.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"retry_upload_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Form Upload", "Password Protected", "Verify upload fails on encrypted password-protected PDF",
        "User is on Upload page.",
        "1. Select password protected PDF.\n2. Tap 'Upload'.",
        "File: locked.pdf",
        "App/Server rejects file: 'PDF is password-protected. Decrypt it before uploading.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'protected')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Form Upload", "Metadata Retention", "Verify uploaded form metadata is recorded in local database",
        "User uploads form successfully.",
        "1. Complete upload flow.\n2. Open Form History.\n3. Verify form entry name and date.",
        "N/A",
        "Local database logs metadata record: Form name, type, size, upload date.",
        "High", "Major", "Appium Native", "accessibilityId(\"history_item_0\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "Invalid Mime Type", "Verify file header verification (Magic Bytes check)",
        "User uploads text file renamed to .png.",
        "1. Select text file renamed to image.png.\n2. Tap 'Upload'.",
        "File: fake_image.png (contains plain text)",
        "Upload fails. Server validation intercepts invalid magic bytes.",
        "High", "Critical", "Appium Native", "xpath(\"//*[contains(@text, 'Invalid')]\")", "Yes",
        ["Security", "Regression", "High Priority Automation"]
    ),
    (
        "Form Upload", "Concurrent Uploads", "Verify app restricts uploading multiple forms concurrently from UI",
        "User starts upload.",
        "1. Tap upload button.\n2. Immediately try tapping other items or re-tapping upload.\n3. Verify state.",
        "N/A",
        "Upload button is disabled, background gestures blocked until process completes.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"upload_btn\")", "Yes",
        ["Regression"]
    ),

    # H. Guidance Testing (20 TCs)
    (
        "Guidance", "Field Detection", "Verify dynamic boxes match layout parameters on document",
        "OCR text extraction completed.",
        "1. Look at screen viewport.\n2. Observe colored outline boxes on input fields.",
        "N/A",
        "Outline boxes overlap input lines (Name, Address, etc.) exactly.",
        "High", "Critical", "Appium Native", "accessibilityId(\"highlight_box_0\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Guidance", "Highlighted Box Validation", "Verify box changes color when tapped",
        "Guidance overlays active.",
        "1. Tap highlighted box for 'First Name'.\n2. Observe border color modification.",
        "N/A",
        "Box color shifts from Blue (inactive) to Green/Yellow (active focus).",
        "High", "Major", "Appium Native", "accessibilityId(\"highlight_box_0\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Guidance", "Guidance Card", "Verify guidance card opens with detail information on tapped box",
        "Guidance overlays active.",
        "1. Tap a highlighted field box.\n2. Observe tooltip or guidance bottom sheet card.",
        "Field: 'Signature'",
        "Bottom sheet pops up. Explains: 'Sign inside box. Ensure signature matches bank records.'",
        "High", "Critical", "Appium Native", "accessibilityId(\"guidance_card_title\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Guidance", "Step Navigation", "Verify next step navigation button details",
        "Guidance bottom card active.",
        "1. Tap 'Next Step' button on guidance card.\n2. Verify highlighted box index changes.",
        "N/A",
        "Card steps to next coordinate sequence. Box highlights next target field.",
        "High", "Major", "Appium Native", "accessibilityId(\"next_step_btn\")", "Yes",
        ["Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Guidance", "Step Navigation", "Verify previous step navigation button details",
        "Guidance bottom card active at Step 2.",
        "1. Tap 'Previous Step' button.\n2. Observe highlight shift.",
        "N/A",
        "Highlight shifts back to Step 1 box.",
        "High", "Major", "Appium Native", "accessibilityId(\"prev_step_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Guidance", "Multi-language Guidance", "Verify translation of guide card to Telugu",
        "Telugu language selected.",
        "1. Tap guide box.\n2. Read instructions inside card.",
        "Language: Telugu",
        "Card content reads in Telugu script correctly.",
        "High", "Major", "Appium Native", "accessibilityId(\"guidance_card_text\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Guidance", "Multi-language Guidance", "Verify translation of guide card to Hindi",
        "Hindi language selected.",
        "1. Tap guide box.\n2. Read instructions inside card.",
        "Language: Hindi",
        "Card content reads in Hindi script correctly.",
        "High", "Major", "Appium Native", "accessibilityId(\"guidance_card_text\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Guidance", "Multi-language Guidance", "Verify translation of guide card to Tamil",
        "Tamil language selected.",
        "1. Tap guide box.\n2. Read instructions.",
        "Language: Tamil",
        "Card content reads in Tamil script.",
        "High", "Major", "Appium Native", "accessibilityId(\"guidance_card_text\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Guidance", "Pinch Zoom Interaction", "Verify guide boxes coordinate rendering on zoom",
        "Guidance overlays active.",
        "1. Double tap screen viewport to zoom in.\n2. Tap highlighted box.",
        "Zoom factor: 2.0x",
        "Target box is tapped correctly, and guidance card displays corresponding help content.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"ocr_viewport\")", "Yes",
        ["Regression"]
    ),
    (
        "Guidance", "Mandatory Field Indicator", "Verify indicator color for mandatory fields",
        "Guidance overlays active.",
        "1. Identify mandatory fields (e.g., Account Number).\n2. Verify border accent.",
        "N/A",
        "Mandatory field boxes have a red asterisk symbol or red border accent indicating criticality.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"highlight_box_required_0\")", "Yes",
        ["Regression"]
    ),
    (
        "Guidance", "Step Count Status", "Verify progress tracker matches current steps indicator",
        "Guidance active on 5-step form.",
        "1. Check status tracker text.\n2. Tap Next Step.",
        "N/A",
        "Initially reads 'Step 1 of 5'. After tapping next, transitions to 'Step 2 of 5'.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"step_count_text\")", "Yes",
        ["Sanity", "Regression"]
    ),
    (
        "Guidance", "Dismiss Guidance", "Verify closing guidance sheet returns to plain form view",
        "Guidance card active.",
        "1. Tap close 'X' button on bottom card or tap empty space.\n2. Verify UI.",
        "N/A",
        "Guidance sheet slides down. Highlights remain or dim appropriately.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"close_guidance_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Guidance", "Help Document", "Verify tooltip button loads details definitions glossary",
        "Guidance card active.",
        "1. Tap small question mark '?' beside label name.\n2. Observe popup.",
        "N/A",
        "Popup box explains banking term (e.g., 'IFSC Code: Indian Financial System Code, 11 chars').",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"term_tooltip_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Guidance", "Voice Toggle Integration", "Verify voice guide starts speaking current highlighted step text",
        "Guidance card active.",
        "1. Tap speaker icon inside card.\n2. Check audio output.",
        "N/A",
        "Voice speaks the exact description text of the highlighted card.",
        "High", "Major", "Appium Native", "accessibilityId(\"card_speak_btn\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Guidance", "E2E Accuracy", "Verify OCR translation mismatch recovery",
        "Low quality OCR output.",
        "1. Navigate guidance.\n2. Verify fallback description overrides garbage values.",
        "N/A",
        "Standard guidance template instructions are displayed even if OCR text has typos.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"guidance_card_text\")", "Yes",
        ["Regression"]
    ),
    (
        "Guidance", "Landscape Redraw", "Verify highlights redraw correctly on screen rotation",
        "Guidance active.",
        "1. Rotate device to landscape layout.\n2. Observe boxes coordinates alignment.",
        "Device rotation",
        "Compose layout recomposes, positioning highlights precisely on the landscape-rotated form coordinates.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"ocr_viewport\")", "Yes",
        ["Regression"]
    ),
    (
        "Guidance", "Scroll View", "Verify auto-scroll focus centering tapped field",
        "Tapping a field box near the bottom of a large document.",
        "1. Tap box at coordinates bottom screen.\n2. Observe viewport viewport.",
        "N/A",
        "Document view auto-scrolls to center the active field box on screen.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"ocr_viewport\")", "Yes",
        ["Regression"]
    ),
    (
        "Guidance", "Form Templates Match", "Verify guidance recognizes custom SBI form template model",
        "Uploaded SBI bank form image.",
        "1. Finish OCR.\n2. Inspect top title indicator.",
        "Form: SBI savings account",
        "App matches template model showing: 'SBI Savings Account Form Guidance' at top.",
        "High", "Major", "Appium Native", "accessibilityId(\"form_template_title\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Guidance", "Form Templates Match", "Verify guidance recognizes Aadhar registration form template",
        "Uploaded Aadhar form.",
        "1. Finish OCR.\n2. Inspect title.",
        "Form: Aadhar card application",
        "Title states: 'Aadhar Card Application Form Guidance'.",
        "High", "Major", "Appium Native", "accessibilityId(\"form_template_title\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Guidance", "Unsupported Template Fallback", "Verify generic guidance loads for unrecognized forms",
        "Uploaded unknown arbitrary form sheet.",
        "1. Run OCR.\n2. Read guidance header.",
        "File: random_bill.jpg",
        "Guidance system switches to OCR Text overlay mode with generic help instructions.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"form_template_title\")", "Yes",
        ["Regression"]
    ),

    # I. Voice Guidance Testing (20 TCs)
    (
        "Voice Guidance", "Audio Generation", "Verify text-to-speech engine loads voice asset on start",
        "User initiates voice guidance mode.",
        "1. Click Voice button.\n2. Listen for initial welcome tone/statement.",
        "N/A",
        "Engine initializes successfully; welcome voice plays: 'FormSahayak voice assistant active.'",
        "High", "Major", "Appium Native", "accessibilityId(\"voice_assistant_toggle\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Voice Guidance", "Audio Playback", "Verify speaking of current guidance step instructions",
        "User is at Step 1 (Name field).",
        "1. Click 'Play Voice'.\n2. Verify headset output.",
        "N/A",
        "Voice assistant speaks clearly: 'Step 1: Enter your full name in block letters as shown in your ID.'",
        "High", "Critical", "Appium Native", "accessibilityId(\"play_voice_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Voice Guidance", "Pause/Resume", "Verify voice pauses on tapping pause button",
        "Voice guide is actively speaking.",
        "1. Tap 'Pause Voice' button.\n2. Verify speaking stops instantly.",
        "N/A",
        "Voice halts speaking. Button icon changes to Play.",
        "High", "Major", "Appium Native", "accessibilityId(\"pause_voice_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Voice Guidance", "Pause/Resume", "Verify voice resumes from exact paused position",
        "Voice guide is paused.",
        "1. Tap 'Resume Voice' (Play icon).\n2. Listen to playback text stream.",
        "N/A",
        "Voice resumes speaking from the exact word it was paused at, rather than restarting.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "Yes",
        ["Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Voice Guidance", "Language Specific Audio", "Verify English voice guide playback",
        "App language set to English.",
        "1. Tap Speak guidance.\n2. Check speech accent and language.",
        "Language: English",
        "Voice guide plays in clear English with neutral/Indian accent.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Voice Guidance", "Language Specific Audio", "Verify Telugu voice guide playback",
        "App language set to Telugu.",
        "1. Tap Speak guidance.\n2. Check speech language.",
        "Language: Telugu",
        "Voice guide speaks fluent Telugu translating the field step guidelines.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Voice Guidance", "Language Specific Audio", "Verify Hindi voice guide playback",
        "App language set to Hindi.",
        "1. Tap Speak guidance.\n2. Check speech language.",
        "Language: Hindi",
        "Voice guide speaks fluent Hindi instructions.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Voice Guidance", "Language Specific Audio", "Verify Tamil voice guide playback",
        "App language set to Tamil.",
        "1. Tap Speak guidance.\n2. Check speech language.",
        "Language: Tamil",
        "Voice guide speaks fluent Tamil instructions.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Voice Guidance", "Volume Controls", "Verify volume shifts via hardware volume buttons",
        "Voice guide active.",
        "1. Press physical Volume Up/Down button on device.\n2. Observe audio level changes.",
        "Hardware press",
        "TTS stream volume conforms immediately with system media channel volume levels.",
        "Medium", "Minor", "Appium Native", "N/A (Hardware button)", "Yes",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Mute Toggle", "Verify audio mute button toggles output volume to zero",
        "Voice guide active.",
        "1. Click 'Mute' speaker icon.\n2. Verify speaking output.",
        "N/A",
        "Audio output cuts out. Progress bar keeps moving silently. Icon displays mute.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"mute_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Speed Settings", "Verify voice playback rate adjustment to 0.75x",
        "Voice guide active.",
        "1. Open voice speed selector.\n2. Choose '0.75x'.\n3. Observe voice rhythm.",
        "Speed: 0.75x",
        "Speaking rate slows down, remaining readable and distinct.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"speed_slow_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Speed Settings", "Verify voice playback rate adjustment to 1.5x",
        "Voice guide active.",
        "1. Open voice speed selector.\n2. Choose '1.5x'.\n3. Observe voice speed.",
        "Speed: 1.5x",
        "Speaking rate speeds up correctly.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"speed_fast_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Headset Unplug", "Verify voice pauses automatically when headphones are unplugged",
        "Voice guide active. Earphones plugged in.",
        "1. Disconnect headphones jack or Bluetooth connection.\n2. Verify app playback status.",
        "Headphones disconnected",
        "App pauses voice playback immediately preventing sound playing on public speakers.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "No",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Phone Call Interruption", "Verify voice pauses on incoming phone call",
        "Voice guide active.",
        "1. Receive a simulated incoming GSM call.\n2. Verify if speech halts.",
        "Incoming Call",
        "Audio guidance pauses immediately. App moves to background state.",
        "High", "Major", "Appium Native", "N/A (System GSM simulation)", "No",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Phone Call Interruption", "Verify voice resumes after ending phone call",
        "Voice guide was paused by incoming call.",
        "1. Terminate phone call.\n2. Resume FormSahayak app.\n3. Check voice status.",
        "Call ended",
        "Voice resumes or prompts user to resume speaking from where it left off.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "No",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Screen Lock Behavior", "Verify voice continues playing when screen locks",
        "Voice guide active.",
        "1. Press Power button to lock screen.\n2. Listen for voice playback.",
        "Screen Locked",
        "TTS stream continues speaking instructions sequentially even with blanked locked display.",
        "Medium", "Minor", "Appium Native", "N/A (Lock screen audio channel)", "No",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Audio File Fallback", "Verify fallback to localized audio tracks when TTS engine fails",
        "TTS engine is disabled/missing on user device.",
        "1. Launch speech guidance on device without Google TTS.\n2. Verify speech output.",
        "TTS missing",
        "App falls back to playing pre-recorded audio dictionary file blocks (.mp3) without errors.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "No",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Auto Step Forward", "Verify voice guide auto-steps to next instruction when current finish",
        "User is listening to end of step 1 guidance.",
        "1. Wait until voice finishes reading step 1 instructions.\n2. Observe app behavior.",
        "N/A",
        "App auto-navigates highlight to Step 2 and begins speaking Step 2 details (if auto-forward enabled).",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"step_count_text\")", "Yes",
        ["Regression"]
    ),
    (
        "Voice Guidance", "Offline Audio Guidance", "Verify voice guidance works when device has no network",
        "Device disconnected from internet.",
        "1. Launch guidance.\n2. Tap speak instructions.\n3. Verify audio feedback.",
        "Network: Offline",
        "Voice plays successfully using on-device cached TTS languages libraries.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Voice Guidance", "Volume Overlay UI", "Verify app displays localized speech progress text on screen",
        "Voice guide active.",
        "1. Observe bottom screen captions bar.\n2. Compare spoken text with captions text.",
        "N/A",
        "Captions display a text script matching spoken words in real time.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"voice_captions_text\")", "Yes",
        ["UI/UX", "Regression"]
    ),

    # J. Form History Testing (15 TCs)
    (
        "Form History", "View History", "Verify previously uploaded forms are listed in Form History screen",
        "User is logged in. Form history has 3 uploaded records.",
        "1. Navigate to 'Form History' tab.\n2. Observe list items.",
        "N/A",
        "All 3 uploaded forms are rendered with metadata (name, date, status).",
        "High", "Critical", "Appium Native", "accessibilityId(\"history_item_0\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Form History", "Search History", "Verify filtering history list by document name search",
        "History contains 'SBI Savings Form', 'PAN card application'.",
        "1. Input 'SBI' into search box.\n2. Tap enter / watch updates.",
        "Search: 'SBI'",
        "List updates instantly, showing only 'SBI Savings Form'. 'PAN card' is hidden.",
        "High", "Major", "Appium Native", "accessibilityId(\"search_history_input\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Form History", "History Persistence", "Verify history remains updated after logout and logging back in",
        "User has uploaded forms history.",
        "1. Check history.\n2. Log out.\n3. Log back in.\n4. Check history again.",
        "N/A",
        "History items remain identical. Loaded from backend database syncing client state.",
        "High", "Critical", "Appium Native", "accessibilityId(\"history_item_0\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Form History", "Empty History State", "Verify empty state screen placeholder illustration",
        "New user account with zero history logs.",
        "1. Navigate to Form History.\n2. Inspect screen layout.",
        "N/A",
        "Screen prints: 'No forms uploaded yet. Tap \"Upload Form\" to begin.' with placeholder icon.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"empty_history_text\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Form History", "Delete Item", "Verify successful removal of a single form record from history list",
        "History lists multiple uploaded forms.",
        "1. Swipe left on history item or tap trash icon.\n2. Tap 'Confirm' on delete dialog.\n3. Observe list.",
        "Delete target: Item index 0",
        "Item disappears. Server deletes file and record. Success toast message displayed.",
        "High", "Major", "Appium Native", "accessibilityId(\"delete_history_btn_0\")", "Yes",
        ["Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Form History", "Scroll Performance", "Verify scroll smoothness of history list with 50+ elements",
        "History contains 50 uploaded forms.",
        "1. Perform rapid vertical scroll sweeps.\n2. Observe UI rendering lags.",
        "Scroll action",
        "Scrolling is smooth. Jetpack Compose LazyColumn recycles views with no lags or crashes.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"history_lazy_list\")", "Yes",
        ["Performance", "Regression"]
    ),
    (
        "Form History", "Search History", "Verify search returns blank message on invalid search query",
        "User has forms list.",
        "1. Type 'xyzgarbage' into search query.\n2. Observe list area.",
        "Search: 'xyzgarbage'",
        "Shows: 'No matching documents found.'",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"search_no_results_text\")", "Yes",
        ["Regression"]
    ),
    (
        "Form History", "Details Link", "Verify tapping history item redirects to its guidance screen",
        "User is in history view.",
        "1. Tap on history card 'SBI Savings Form'.\n2. Observe target screen.",
        "N/A",
        "Opens document view directly displaying the pre-processed OCR coordinate overlays and steps.",
        "High", "Major", "Appium Native", "accessibilityId(\"history_item_0\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Form History", "Search History", "Verify search query clear button resets the list view",
        "User searched for 'SBI'.",
        "1. Click the 'X' button inside search textbox.\n2. Verify list display.",
        "N/A",
        "Search text is cleared, full list of history items displays.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"clear_search_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Form History", "Offline Access", "Verify accessing offline cached forms list when offline",
        "User is offline. Has previous local database rows.",
        "1. Set phone to Airplane mode.\n2. Navigate to Form History.",
        "N/A",
        "Cached forms list renders from SQLite database with alert: 'Showing local cached list. Upload is disabled.'",
        "High", "Major", "Appium Native", "accessibilityId(\"history_offline_alert\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Form History", "Sort History", "Verify sorting history items by date",
        "History contains documents uploaded on different dates.",
        "1. Click Sort filter.\n2. Select 'Date: Newest First'.\n3. Observe sequence.",
        "Sort: Newest First",
        "Items sort in chronological descending order.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"sort_date_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Form History", "Sort History", "Verify sorting history items alphabetically by name",
        "Multiple items present.",
        "1. Click Sort filter.\n2. Select 'Name: A to Z'.",
        "Sort: Name A-Z",
        "Items sequence updates alphabetically.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"sort_name_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Form History", "Image Thumbnails", "Verify thumbnail images load successfully using Coil library",
        "Multiple history items list.",
        "1. Inspect card layouts.\n2. Verify thumbnail image load status.",
        "N/A",
        "Small preview of document loads inside each card avatar circle successfully.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"history_thumbnail_0\")", "Yes",
        ["Regression"]
    ),
    (
        "Form History", "Delete Item", "Verify cancel action on deletion confirmation prompt",
        "User click delete.",
        "1. Tap delete trash icon.\n2. Alert displays. Tap 'Cancel'.\n3. Check list status.",
        "N/A",
        "Confirmation dialog is closed. Form item is not deleted.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"confirm_delete_cancel\")", "Yes",
        ["Regression"]
    ),
    (
        "Form History", "Local Cache Sync", "Verify cache updates when adding a new form offline and then going online",
        "Offline sync active.",
        "1. Upload form offline (cached locally).\n2. Connect to internet.\n3. Check if server receives sync request.",
        "N/A",
        "Metadata synchronizes to Railway cloud DB and updates history server timestamp.",
        "High", "Major", "Appium Native", "accessibilityId(\"history_item_0\")", "No",
        ["Regression"]
    ),

    # K. Feedback Testing (15 TCs)
    (
        "Feedback", "Submit Feedback", "Verify submitting valid feedback with rating and text message",
        "User is logged in. Navigates to Feedback screen.",
        "1. Select 4 stars on rating bar.\n2. Write feedback text.\n3. Tap 'Submit'.",
        "Rating: 4 Stars, Text: 'App OCR translates Telugu perfectly. Awesome guidance.'",
        "Feedback uploaded. Success dialog shows: 'Thank you for your feedback!'. Directed to dashboard.",
        "High", "Major", "Appium Native", "accessibilityId(\"submit_feedback_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "Feedback", "Rating Validation", "Verify submission fails if no star rating is selected",
        "User is on Feedback screen.",
        "1. Leave stars empty (0 selected).\n2. Input text message.\n3. Tap 'Submit'.",
        "Rating: 0 Stars, Text: 'Good app'",
        "Error message displayed: 'Please select a rating score.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'select a rating')]\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Feedback", "Empty Feedback", "Verify submission fails if comments text box is empty",
        "User is on Feedback screen.",
        "1. Select 5 stars.\n2. Leave comment box blank.\n3. Tap 'Submit'.",
        "Rating: 5 Stars, Text: ''",
        "Validation error shows: 'Please write a brief comment.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'write a brief comment')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Feedback", "Feedback Limits", "Verify comment box character limit constraints (max 500 characters)",
        "User types extremely long comment.",
        "1. Paste block of text containing 550 characters into comment input.\n2. Tap 'Submit' or observe textbox counter.",
        "Text: 550 characters input",
        "Text is truncated at 500 characters, or error shows: 'Maximum length exceeded.'",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"feedback_char_counter\")", "Yes",
        ["Regression"]
    ),
    (
        "Feedback", "Multiple Submissions", "Verify user is blocked from rapid duplicate submissions",
        "User already submitted feedback once.",
        "1. Attempt navigating to feedback page again.\n2. Observe screen state or error alert.",
        "N/A",
        "App blocks submitting again: 'Feedback already submitted for this version.' or allows editing old rating.",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'already submitted')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Feedback", "Offline Submission Queue", "Verify feedback is cached and queued if submitted when offline",
        "Device network is disconnected.",
        "1. Write feedback.\n2. Click 'Submit'.\n3. Connect network.",
        "Network State: Offline",
        "Message indicates: 'Offline. Feedback saved locally and will upload when online.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'saved locally')]\")", "No",
        ["Regression"]
    ),
    (
        "Feedback", "Rating UI Verification", "Verify tapping star highlights preceding stars",
        "User is on feedback page.",
        "1. Tap star rating index 3.\n2. Observe stars index 1 and 2.",
        "N/A",
        "Stars 1, 2, and 3 light up in accent color (e.g. Amber). Stars 4 and 5 remain grey.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"star_rating_3\")", "Yes",
        ["UI/UX", "Regression"]
    ),
    (
        "Feedback", "Input Sanitization", "Verify SQL injection check in comments text box",
        "User is on feedback page.",
        "1. Enter SQL query inside comments.\n2. Tap 'Submit'.",
        "Comments: \"'; DROP TABLE feedback;--\"",
        "Server sanitizes comment input. Entry saved literally. No DB drops occur.",
        "High", "Critical", "Appium Native", "accessibilityId(\"submit_feedback_btn\")", "Yes",
        ["Security", "Regression", "High Priority Automation"]
    ),
    (
        "Feedback", "Input Sanitization", "Verify XSS scripting in comments text box",
        "User is on feedback screen.",
        "1. Input Javascript prompt scripts into comments box.\n2. Submit feedback.\n3. Open feedback reviews log in Admin Panel.",
        "Comments: \"<svg onload=alert(1)>\"",
        "Admin browser does not execute script; strings are encoded and rendered textually.",
        "High", "Critical", "Appium Native", "accessibilityId(\"submit_feedback_btn\")", "Yes",
        ["Security", "Regression", "High Priority Automation"]
    ),
    (
        "Feedback", "Special Characters", "Verify comment box handles emojis and native scripts",
        "User submits feedback with emojis and Telugu script.",
        "1. Enter comments.\n2. Tap 'Submit'.",
        "Comments: 'చాలా మంచి యాప్! 😃👍'",
        "Feedback stores successfully. DB supports UTF-8 character coding.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"submit_feedback_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Feedback", "Cancel Action", "Verify back button closes feedback dialog without saving",
        "Feedback dialog prompted.",
        "1. Modify text content.\n2. Press back hardware button.",
        "N/A",
        "Dialog closes. No feedback upload API is initiated. Old page loads.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"feedback_back_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Feedback", "Submit Feedback", "Verify feedback status is updated in Railway MySQL cloud database",
        "User completes feedback submission.",
        "1. Query MySQL feedback table for user ID.\n2. Verify row values.",
        "MySQL select query",
        "Row values match precisely with ratings score and text submitted.",
        "High", "Major", "Appium Native", "N/A (DB query)", "No",
        ["Regression"]
    ),
    (
        "Feedback", "Network Error Handling", "Verify spinner ceases and displays error on API failure during feedback",
        "Backend mock server returns error 500 on feedback post.",
        "1. Enter feedback inputs.\n2. Click 'Submit'.",
        "N/A",
        "Spinner terminates. Warning displays: 'Failed to submit feedback. Try again later.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Failed to submit')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Feedback", "Rating UI Verification", "Verify sliding gesture on rating bar updates score",
        "User slides thumb across rating stars.",
        "1. Slide thumb from star 1 to star 5.\n2. Check rating score visual changes.",
        "Slide gesture",
        "Stars count lights up incrementally tracking swipe position.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"rating_bar\")", "No",
        ["UI/UX", "Regression"]
    ),
    (
        "Feedback", "Login Validation", "Verify unauthorized guests cannot access feedback submission APIs",
        "User is not logged in.",
        "1. Trigger POST `/api/feedback` via external curl tool.\n2. Check return code.",
        "Payload: {rating:5}",
        "Response status code 401 Unauthorized is returned.",
        "High", "Critical", "Appium Native", "N/A (API validation)", "No",
        ["Security", "Regression"]
    ),

    # L. Network Testing (15 TCs)
    (
        "Network", "Slow Network", "Verify app displays loader during API calls on 3G slow connection",
        "Emulating slow network connection constraints (3G slow: 300 Kbps).",
        "1. Click 'Run OCR' on document.\n2. Verify presence of loader animation.",
        "Bandwidth constraint 300 Kbps",
        "Loading spinner overlay appears immediately, preventing UI interaction until OCR completes.",
        "High", "Major", "Appium Native", "accessibilityId(\"ocr_loading_spinner\")", "Yes",
        ["Smoke", "Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Network", "No Internet", "Verify app behavior when internet cuts out during PDF upload",
        "Uploading document.",
        "1. Tap 'Upload'.\n2. Immediately cut network connection.\n3. Check error notifications.",
        "N/A",
        "App detects network failure. Stops upload, shows retry overlay: 'Network lost. Check settings.'",
        "High", "Critical", "Appium Native", "xpath(\"//*[contains(@text, 'Network lost')]\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Network", "API Timeout", "Verify client handling on backend response delay timeout (30 seconds)",
        "API server simulated hung or extremely delayed.",
        "1. Click 'Profile' save change.\n2. Wait for app response limits.",
        "Delay simulation: 35s",
        "OkHttp client triggers timeout exception after 30s. App shows: 'Request timed out. Please try again.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'timed out')]\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Network", "Backend Unavailable", "Verify error handling when Render hosting backend server is down (503)",
        "Render web server simulated crashed or offline.",
        "1. Launch application.\n2. Click Login.\n3. Check error display.",
        "Server status: 503 Service Unavailable",
        "App displays error toast: 'Server is currently undergoing maintenance. Please try again later.'",
        "High", "Critical", "Appium Native", "xpath(\"//*[contains(@text, 'maintenance')]\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Network", "Connection Switch", "Verify seamless operation during network switch (Wi-Fi to Mobile Data)",
        "OCR is in progress state.",
        "1. Disconnect Wi-Fi.\n2. Device switches to cellular data.\n3. Check API processing.",
        "N/A",
        "Connection switches gracefully. Client socket automatically retries request; transaction completes.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"ocr_overlay\")", "No",
        ["Regression"]
    ),
    (
        "Network", "Offline Launch", "Verify app launches gracefully in Offline mode",
        "Device is in airplane mode.",
        "1. Launch FormSahayak.\n2. Verify app does not crash or loop on splash.",
        "Airplane mode",
        "App loads homepage dashboard, showing offline status indicator at top: 'Offline Mode - Limited Functions'.",
        "High", "Major", "Appium Native", "accessibilityId(\"offline_status_indicator\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Network", "Offline Launch", "Verify login is blocked during offline state",
        "Device is offline.",
        "1. Input valid credentials on login screen.\n2. Tap 'Login'.",
        "N/A",
        "App intercepts submit, showing: 'Login requires active network connection.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'requires active network')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Network", "Offline Guidance", "Verify cached guidance templates load without internet",
        "Device is offline. User has already cached form configurations.",
        "1. Open details of cached Form History item.\n2. Observe layout overlays.",
        "N/A",
        "Guidance overlays load successfully from SQLite database locally.",
        "High", "Major", "Appium Native", "accessibilityId(\"ocr_overlay\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Network", "Offline Profile Edit", "Verify profile editing fails offline with prompt warning",
        "Device offline.",
        "1. Edit name in profile settings.\n2. Click 'Save'.",
        "N/A",
        "Changes blocked. Alerts: 'Network required to save profile updates.'",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'Network required')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Network", "Offline Feedback", "Verify offline feedback warning message",
        "Device offline.",
        "1. Complete rating and comment field.\n2. Click 'Submit'.",
        "N/A",
        "Feedback is saved locally in background queue: 'Saved. Feedback will upload when network is available.'",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'Saved')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Network", "DNS Resolution Fail", "Verify client handling when API domain name cannot be resolved",
        "Client DNS returns invalid routing.",
        "1. Trigger API operation.\n2. Verify client handler response.",
        "DNS resolution failed",
        "Handles UnknownHostException. Alerts user: 'Connection error. Check device settings.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Connection error')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Network", "SSL Handshake Fail", "Verify client handling when SSL certificate is invalid",
        "Simulating invalid/expired SSL certificate on backend API.",
        "1. Initiate API request.\n2. Check connection output.",
        "Expired SSL cert",
        "App blocks connection immediately (security policy). Shows: 'Secure connection failed.'",
        "High", "Critical", "Appium Native", "xpath(\"//*[contains(@text, 'Secure connection failed')]\")", "Yes",
        ["Security", "Regression", "High Priority Automation"]
    ),
    (
        "Network", "API Response Error", "Verify app behavior when API returns 500 Internal Server Error",
        "Backend server encounters database script fault during upload details processing.",
        "1. Trigger database operation.\n2. Server responds 500 code.",
        "N/A",
        "App alerts: 'An internal server error occurred. Please try again later.'",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'internal server error')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Network", "Offline Camera", "Verify camera image capturing continues to work offline",
        "Device offline.",
        "1. Tap 'Upload Form' -> 'Take Photo'.\n2. Snap picture.",
        "N/A",
        "Photo is captured, saved to local cache successfully, displaying step: 'Saved locally. Upload pending network.'",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"captured_preview_img\")", "Yes",
        ["Regression"]
    ),
    (
        "Network", "API Response Error", "Verify app handles 404 Not Found response codes safely",
        "App requests non-existent template ID.",
        "1. Request details for non-existing document key.\n2. Watch UI output.",
        "ID: 9999",
        "App handles 404 safely: 'Form template details not found.'",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'not found')]\")", "Yes",
        ["Regression"]
    ),

    # M. Security Testing (20 TCs)
    (
        "Security", "Unauthorized Access", "Verify unauthorized API requests are blocked on backend server",
        "User is not authenticated (missing/invalid JWT token).",
        "1. Make raw HTTP request to `/api/documents` using curl.\n2. Verify response headers and status.",
        "Header: Missing Bearer Token",
        "Returns HTTP 401 Unauthorized status, blocking resource access.",
        "High", "Critical", "Appium Native", "N/A (Curl validation)", "No",
        ["Smoke", "Regression", "Critical Path"]
    ),
    (
        "Security", "Token Validation", "Verify session token is invalidated on the server after logging out",
        "User has logged out from device.",
        "1. Grab the active JWT token before logout.\n2. Log out.\n3. Make raw API request using the copied JWT token.",
        "Header: Authorization Bearer [Old JWT]",
        "Server rejects token with HTTP 401 Unauthorized since token is blacklisted/invalidated.",
        "High", "Critical", "Appium Native", "N/A (Postman validation)", "No",
        ["Regression", "Critical Path"]
    ),
    (
        "Security", "Session Expiration", "Verify token expiration invalidates local active state after 24 hours",
        "User session is active, token holds 24 hour expiry payload.",
        "1. Adjust client system clock forward by 25 hours.\n2. Perform API action inside application.",
        "System time pushed ahead 25h",
        "App immediately intercepts token expiry. Logs user out, redirecting to login with message 'Session expired.'",
        "High", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Session expired')]\")", "No",
        ["Regression"]
    ),
    (
        "Security", "Sensitive Data Protection", "Verify passwords are not visible in plain text inside local logcat streams",
        "User registration or login is running.",
        "1. Enter credentials.\n2. Read Android logcat console stream.\n3. Look for password strings.",
        "Logcat streams",
        "Zero password information is leaked in logs. Logs mask inputs or log generic states.",
        "High", "Critical", "Appium Native", "N/A (Logcat audit)", "No",
        ["Security", "Regression"]
    ),
    (
        "Security", "Sensitive Data Protection", "Verify sensitive text fields block copying (clipboard restriction)",
        "User is typing password.",
        "1. Focus password input box.\n2. Long press cursor.\n3. Verify presence of Copy button.",
        "Password active",
        "Clipboard operations (Copy/Cut) are disabled for password inputs.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"login_password\")", "Yes",
        ["Regression"]
    ),
    (
        "Security", "SQL Injection", "Verify search input text sanitizes SQL wildcards (% and _)",
        "User in search history screen.",
        "1. Input '%' in search box.\n2. Observe returned records.",
        "Query: '%'",
        "No database query crash; treats search as string literal or filters query safely.",
        "High", "Major", "Appium Native", "accessibilityId(\"search_history_input\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Security", "XSS Prevention", "Verify markdown tags inside translation guides do not run scripting payloads",
        "Backend sends malformed translation containing JS tags.",
        "1. Modify Telugu language template details on server to contain alert scripting.\n2. Load guidance page in Telugu.",
        "Payload: <script>alert(1)</script>",
        "App processes translation text as raw string data in Compose Text component; does not execute code.",
        "High", "Critical", "Appium Native", "accessibilityId(\"guidance_card_text\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Security", "Screenshot Block", "Verify screen capture blocking configuration on sensitive screens",
        "User is on profile details screen holding password.",
        "1. Try to take screenshot using hardware keys.\n2. Attempt ADB screen recording.",
        "Screenshot request",
        "OS blocks capture showing: 'Cannot capture screenshot due to security policy.'",
        "Medium", "Major", "Appium Native", "N/A (ADB screen check)", "No",
        ["Regression"]
    ),
    (
        "Security", "Root Detection", "Verify app displays warnings or exits safely on rooted devices",
        "Device environment is rooted.",
        "1. Launch application on rooted emulator/phone.\n2. Observe alerts.",
        "Rooted system",
        "App detects root binaries (su, busybox) and flags warning message or restricts sensitive tasks.",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Root') or contains(@text, 'Security warning')]\")", "No",
        ["Regression"]
    ),
    (
        "Security", "Certificate Pinning", "Verify TLS connection is blocked if server cert is spoofed by proxy (MITM proxy test)",
        "Client is configured with Charles Proxy/Mitmproxy custom certificate.",
        "1. Launch application.\n2. Trigger API call.\n3. Check proxy details.",
        "Proxy certificate active",
        "App throws SSLHandshakeException immediately. Request does not process.",
        "High", "Critical", "Appium Native", "N/A (Proxy validation)", "No",
        ["Regression"]
    ),
    (
        "Security", "Obfuscation Validation", "Verify APK code is obfuscated with ProGuard/R8",
        "Production build APK generated.",
        "1. Decompile APK using JADX tool.\n2. Inspect package classes and symbols.",
        "N/A",
        "Class, method, and variable names are obfuscated (a, b, c). Kotlin source metadata is stripped.",
        "High", "Critical", "Appium Native", "N/A (Decompiler check)", "No",
        ["Regression"]
    ),
    (
        "Security", "Data Encryption", "Verify local database SQLite file is encrypted (SQLCipher check)",
        "Local database is written.",
        "1. Pull sqlite file via ADB.\n2. Attempt opening using DB Browser for SQLite without key.",
        "Database: formsahayak.db",
        "File is unreadable, displaying: 'File is encrypted or is not a database.'",
        "High", "Critical", "Appium Native", "N/A (SQLite file audit)", "No",
        ["Regression"]
    ),
    (
        "Security", "Unauthorized Access", "Verify accessing private endpoints fails with 403 Forbidden on wrong user scope",
        "Regular user logs in.",
        "1. Use token to call Admin dashboard API endpoints `/api/admin/metrics`.\n2. Verify result.",
        "Header: Authorization Bearer [User JWT]",
        "Server rejects request returning HTTP 403 Forbidden code.",
        "High", "Critical", "Appium Native", "N/A (API validation)", "No",
        ["Regression"]
    ),
    (
        "Security", "Sensitive Data Protection", "Verify login details are wiped from RAM after application termination",
        "App process ended.",
        "1. Monitor device RAM dump files.\n2. Look for cached variables.",
        "Memory dump",
        "Variables hold no plain text password residual traces.",
        "Medium", "Minor", "Appium Native", "N/A (RAM profiling)", "No",
        ["Regression"]
    ),
    (
        "Security", "Brute Force Protection", "Verify rate limits on OTP Verification APIs",
        "User is updating phone.",
        "1. Click verify OTP.\n2. Submit mismatching OTP codes 10 times in rapid sequence.",
        "N/A",
        "API blocks IP/account temporary, returning HTTP 429 Too Many Requests.",
        "Medium", "Major", "Appium Native", "xpath(\"//*[contains(@text, 'Too many requests')]\")", "Yes",
        ["Regression"]
    ),
    (
        "Security", "Password Validation", "Verify password strength validation rules are active in backend APIs",
        "User tries submitting registration using simple password directly via HTTP POST.",
        "1. Fire registration request containing password '123'.\n2. Inspect response status.",
        "Password: '123'",
        "Server rejects request with HTTP 422 Unprocessable Entity, detailing validation issues.",
        "High", "Major", "Appium Native", "N/A (API testing)", "No",
        ["Regression"]
    ),
    (
        "Security", "Sensitive Data Protection", "Verify private file storage permissions configuration on Android",
        "Application creates temporary images during crop.",
        "1. Check files under context.getExternalFilesDir().\n2. Verify permissions mask.",
        "N/A",
        "Permissions are set strictly to owner (private directory). Other applications cannot access files.",
        "High", "Major", "Appium Native", "N/A (ADB shell check)", "No",
        ["Regression"]
    ),
    (
        "Security", "Token Validation", "Verify app enforces HTTPS protocol on all API endpoints",
        "App initiates communication.",
        "1. Intercept network packets.\n2. Verify protocol scheme.",
        "Target URL",
        "All requests strictly target https://. Non-secure http:// endpoints are rejected.",
        "High", "Critical", "Appium Native", "N/A (Network capturing)", "No",
        ["Regression"]
    ),
    (
        "Security", "SQL Injection", "Verify user feedback inputs sanitize quote character injections",
        "User is in feedback comment box.",
        "1. Enter single quote ('), double quote (\"), and semicolon (;).\n2. Submit comments.",
        "Comments: '\";--'",
        "Comments save successfully. Characters are sanitized correctly to prevent SQL escapes.",
        "High", "Major", "Appium Native", "accessibilityId(\"submit_feedback_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Security", "XSS Prevention", "Verify input sanitation on all API JSON string request payloads",
        "Sending feedback programmatically.",
        "1. Fire POST request with JSON string containing HTML tags.\n2. Verify database records.",
        "JSON payload: {comment: '<a>test</a>'}",
        "Payload is sanitized and stored without executable formatting script tags.",
        "High", "Major", "Appium Native", "N/A (API validation)", "No",
        ["Regression"]
    ),

    # N. UI/UX Testing (25 TCs)
    (
        "UI/UX", "Buttons", "Verify button click state triggers material ripple feedback visual indicator",
        "User is on the Dashboard screen.",
        "1. Touch and hold the 'Upload Form' button.\n2. Observe visual transition.",
        "N/A",
        "Material design ripple effect animates on button background during tap.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"upload_form_btn\")", "Yes",
        ["Sanity", "Regression"]
    ),
    (
        "UI/UX", "Navigation", "Verify switching dashboard screens using bottom navigation tabs",
        "User is on the Dashboard screen.",
        "1. Tap 'Form History' navigation tab.\n2. Tap 'Profile' navigation tab.\n3. Tap 'Home' navigation tab.",
        "N/A",
        "Views transition immediately. Active tab indicator updates to focus matching selection.",
        "High", "Critical", "Appium Native", "accessibilityId(\"navigation_history_tab\")", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "UI/UX", "Screen Orientation", "Verify dashboard view layout preserves alignment on screen rotation",
        "User is on the Dashboard screen.",
        "1. Rotate device landscape orientation.\n2. Check grid layout alignment.\n3. Rotate back to portrait.",
        "N/A",
        "Layout adapts cleanly. Scroll view triggers automatically if height constraints require it.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_scroll_view\")", "Yes",
        ["Sanity", "Regression"]
    ),
    (
        "UI/UX", "Dark Mode", "Verify UI theme matches system dark mode configurations automatically",
        "Device system theme is configured to Dark Mode.",
        "1. Launch FormSahayak application.\n2. Observe background colors and text contrasts.",
        "System Theme: Dark",
        "App loads Dark theme palette (e.g. dark grey backgrounds, white text). Text remains perfectly legible.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Sanity", "Regression"]
    ),
    (
        "UI/UX", "Accessibility", "Verify accessibility screen-reader content descriptions for all icons",
        "TalkBack screen-reader service enabled.",
        "1. Focus cursor on voice guidance mic icon.\n2. Listen for voice-over text output.",
        "TalkBack ON",
        "TalkBack speaks description clearly: 'Voice guidance, button. Double tap to activate.'",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"voice_mic_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Keyboard Dismissal", "Verify keyboard disappears on tapping background canvas area",
        "Keyboard is active on input field.",
        "1. Tap empty background screen canvas area.\n2. Observe keyboard status.",
        "Keyboard displayed",
        "Keyboard hides immediately. Input field loses active focus.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"login_background_canvas\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Language Renders", "Verify Telugu script character rendering boundaries",
        "Telugu language selected.",
        "1. Check labels text boundaries.\n2. Verify characters do not clip on buttons.",
        "Language: Telugu",
        "Telugu text wraps cleanly. Line heights render correctly, showing no overlapping script parts.",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[@text='స్వాగతం']\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Language Renders", "Verify Hindi script character rendering boundaries",
        "Hindi language selected.",
        "1. Check labels text boundaries.\n2. Verify characters do not clip.",
        "Language: Hindi",
        "Hindi characters display cleanly with correct fonts and alignments.",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[@text='स्वागत']\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Language Renders", "Verify Tamil script character rendering boundaries",
        "Tamil language selected.",
        "1. Check labels text boundaries.\n2. Verify characters do not clip.",
        "Language: Tamil",
        "Tamil characters display cleanly.",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[@text='வரவேற்பு']\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Scroll Views", "Verify profile scroll view boundaries on small screens",
        "Small 4-inch display emulator active.",
        "1. Navigate Profile edit.\n2. Scroll down details list.",
        "Screen size: 4-inch",
        "Scroll view scrolls smoothly. Inputs are easily reachable and not hidden under buttons.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"profile_scroll_view\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Buttons", "Verify buttons are disabled during active API processes",
        "OCR processing active.",
        "1. Attempt to tap 'Upload Form' button or back navigation arrow.\n2. Check status.",
        "N/A",
        "Buttons display grayed-out disabled states. Inputs are locked, preventing duplicate actions.",
        "High", "Major", "Appium Native", "accessibilityId(\"upload_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "UI/UX", "Status Bar", "Verify status bar color matches app primary branding colors",
        "App launch completed.",
        "1. Observe status bar color at very top of phone display.",
        "N/A",
        "Status bar theme is color-matched with Compose primary topbar theme (e.g. Steel blue background, white text).",
        "Low", "Minor", "Appium Native", "N/A (Visual status check)", "No",
        ["Regression"]
    ),
    (
        "UI/UX", "Toast Alerts", "Verify error toast disappears automatically after 3.5 seconds",
        "Error occurs (e.g. upload fail).",
        "1. Trigger error warning toast.\n2. Measure displaying duration.",
        "N/A",
        "Toast message is displayed clearly and then fades out after exactly 3.5 seconds.",
        "Low", "Minor", "Appium Native", "xpath(\"//*[contains(@text, 'Failed')]\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "About Page", "Verify 'About Developer' page links open correct socials",
        "About Developer page active.",
        "1. Tap developer GitHub profile link.\n2. Observe redirection.",
        "GitHub Link",
        "Device web browser launches loading correct Github address.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"github_profile_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Language Selection", "Verify selection indicator checkmark displays beside active language",
        "Language selection modal active.",
        "1. Tap Hindi language row.\n2. Observe indicator tick mark.",
        "N/A",
        "Tick checkmark icon appears instantly beside Hindi row, disappears from English row.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"hindi_lang_tick\")", "Yes",
        ["Sanity", "Regression"]
    ),
    (
        "UI/UX", "Back Nav Stack", "Verify back arrow on toolbar pops screen from navigation stack",
        "User is inside history details screen.",
        "1. Tap back arrow icon on top header toolbar.\n2. Observe active view.",
        "N/A",
        "Screen closes. App returns to the previous History list screen view.",
        "High", "Major", "Appium Native", "accessibilityId(\"toolbar_back_btn\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "UI/UX", "Coil Placeholders", "Verify Coil image loading error placeholders display on missing files",
        "Network connection cut, profile avatar image fails to download.",
        "1. Open profile page.\n2. Observe avatar image frame.",
        "N/A",
        "App shows fallback vector image placeholder in place of actual photo without crashing.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"profile_avatar_placeholder\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Font Scaling", "Verify text layouts when Android system font size is set to Largest",
        "Android settings -> Display -> Font Size -> set to Largest.",
        "1. Open FormSahayak dashboard.\n2. Inspect labels text overlap.",
        "Font Size: Max",
        "Labels adjust size or auto-wrap safely without clipping or rendering text blocks outside container.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Keyboard Padding", "Verify scroll behavior of input fields when soft keyboard pops",
        "User focuses on confirm password registration field.",
        "1. Tap input textbox.\n2. Keyboard expands.\n3. Observe text field focus placement.",
        "N/A",
        "Compose view auto-scrolls to push the active input field above the virtual keyboard display area.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"confirm_password_input\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Theme Toggle", "Verify manual light/dark mode switch changes theme instantly",
        "User on Profile Settings screen.",
        "1. Click toggle switch for Dark Mode theme.\n2. Toggle it off.",
        "N/A",
        "App colors shift dynamically between dark theme and light theme formats instantly without page reload.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dark_mode_switch\")", "Yes",
        ["Sanity", "Regression"]
    ),
    (
        "UI/UX", "Splash Animation", "Verify splash screen displays branding logo clean configuration",
        "App launching.",
        "1. Launch app.\n2. Observe splash graphics logo.",
        "N/A",
        "Logo is scaled correctly, holds correct layout proportions and shows no rendering issues.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"splash_logo\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Feedback Stars Layout", "Verify tap coordinates target size on feedback stars selection",
        "User is on feedback view.",
        "1. Select ratings using edge positions tap of the stars.\n2. Verify input response.",
        "N/A",
        "Stars detect clicks efficiently. Touch targets hold minimum dimensions of 48dp x 48dp.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"star_rating_4\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Tab Swiping Gestures", "Verify swiping horizontally transitions tabs in dashboard view",
        "User is on Home Dashboard.",
        "1. Swipe finger horizontally from right to left.\n2. Check view changes.",
        "Swipe gesture",
        "Dashboard transitions smoothly to history tab.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_viewpager\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Card Elevations", "Verify shadow elevations on Compose cards",
        "App dashboard open.",
        "1. Observe elevations shadows surrounding dashboard category selector cards.",
        "N/A",
        "Card borders display subtle distinct shadows rendering card overlays apart from backgrounds.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"home_card_upload\")", "Yes",
        ["Regression"]
    ),
    (
        "UI/UX", "Text Field Errors", "Verify helper error label colors match error design specifications",
        "Registration validation errors active.",
        "1. Trigger error validation.\n2. Check color parameters of error text labels.",
        "N/A",
        "Error texts highlight in specific designated Red color accent (e.g. Material Theme Error).",
        "Low", "Minor", "Appium Native", "xpath(\"//*[@text='Username is required.']\")", "Yes",
        ["Regression"]
    ),

    # O. Performance Testing (15 TCs)
    (
        "Performance", "App Startup Time", "Verify cold launch duration limits to dashboard screen",
        "Application processes are cleared from background.",
        "1. Click app launcher icon.\n2. Stop timer when dashboard elements render.",
        "N/A",
        "Render time must remain strictly under 2.0 seconds.",
        "High", "Major", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Performance", "App Startup Time", "Verify warm launch duration limits",
        "Application is in background.",
        "1. Tap app icon.\n2. Stop timer on display return.",
        "N/A",
        "Warm launch takes under 0.8 seconds to display screen.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Regression"]
    ),
    (
        "Performance", "OCR Processing Time", "Verify EasyOCR API execution and render limit under 10 seconds",
        "Normal quality JPG uploaded.",
        "1. Click 'Run OCR'.\n2. Track duration until coordinates render.",
        "File: test_doc.jpg",
        "Execution completes and rendering highlights display in under 10.0 seconds.",
        "High", "Major", "Appium Native", "accessibilityId(\"ocr_overlay\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Performance", "Large Image Handling", "Verify image memory utilization under high resolution upload",
        "Coil loading active on 4.5MB image template.",
        "1. Open details view.\n2. Monitor memory footprint using android profiling tools.",
        "File: heavy_template.png",
        "Memory allocation remains stable, throwing no OutOfMemory (OOM) exceptions.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"ocr_viewport\")", "No",
        ["Performance", "Regression"]
    ),
    (
        "Performance", "Concurrent Operations", "Verify app behavior when voice plays during OCR processing",
        "OCR active in background.",
        "1. Initiate speech guidance.\n2. Verify CPU usage and voice stuttering status.",
        "N/A",
        "Voice speaks smoothly. Rendering frames do not stutter (FPS remains above 55).",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"play_voice_btn\")", "No",
        ["Performance", "Regression"]
    ),
    (
        "Performance", "Battery Drain", "Verify battery consumption remains low under 10 minutes voice playback",
        "Voice guidance active.",
        "1. Run voice assistant speech continuously for 10 minutes.\n2. Track battery drop percentage.",
        "10 mins audio",
        "Battery drain level is under 1% of total capacity.",
        "Low", "Minor", "Appium Native", "N/A (Battery check)", "No",
        ["Regression"]
    ),
    (
        "Performance", "App Size Growth", "Verify application cached file growth remains bounded",
        "App used extensively to upload forms for 1 week.",
        "1. Check data folder directories metrics.\n2. Observe cache usage.",
        "N/A",
        "Image disk cache is bounded to 100MB limit. Auto-prunes oldest items.",
        "Low", "Minor", "Appium Native", "N/A (ADB storage shell)", "No",
        ["Regression"]
    ),
    (
        "Performance", "Scale History DB", "Verify history database query speed with 1000 items",
        "Local SQLite database holds 1000 mocked metadata records.",
        "1. Load Form History tab.\n2. Measure query load delay.",
        "1000 records",
        "List renders items in under 150ms. Employs cursor paginated loading.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"history_lazy_list\")", "Yes",
        ["Performance", "Regression"]
    ),
    (
        "Performance", "UI Frame Drops", "Verify scroll list performance metrics for zero UI frame drops",
        "History LazyColumn actively scrolling.",
        "1. Scroll list.\n2. Read GPU profiling frames layout logs.",
        "N/A",
        "Visual list displays 0 janky frames. Stays above 60 FPS standard rates.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"history_lazy_list\")", "No",
        ["Regression"]
    ),
    (
        "Performance", "Network Data Usage", "Verify image compression before upload to conserve data",
        "4MB raw photo captured by camera.",
        "1. Tap 'Upload Form'.\n2. Monitor actual payload size sent over network.",
        "File: 4MB raw JPG",
        "App compresses image client-side to under 800KB before transmission, conserving user cellular data.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"upload_btn\")", "No",
        ["Regression"]
    ),
    (
        "Performance", "CPU Spike Limits", "Verify CPU usage spikes do not exceed 40% during local ML Kit scan",
        "ML Kit text parsing active on camera view.",
        "1. Scan doc layout using on-device ML Kit camera detector.\n2. Monitor CPU usage.",
        "N/A",
        "CPU utilization spikes peak below 40% threshold. Device does not heat up.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"camera_preview\")", "No",
        ["Regression"]
    ),
    (
        "Performance", "SQLite Transaction Latency", "Verify local DB write transactions take under 50ms",
        "Saving form history metadata.",
        "1. Upload form.\n2. Monitor DB transaction logs.",
        "N/A",
        "Write transactions to SQLite complete in under 50ms.",
        "Low", "Minor", "Appium Native", "N/A (DB write log check)", "No",
        ["Regression"]
    ),
    (
        "Performance", "Leak Canary Validation", "Verify memory leak check runs cleanly on Profile transitions",
        "Navigating between Profile details, Edit profile, and Home screens multiple times.",
        "1. Execute page jumps 20 times.\n2. Inspect LeakCanary alerts log.",
        "N/A",
        "Zero memory leaks flagged. Component scopes cleaned on destroy.",
        "High", "Major", "Appium Native", "N/A (LeakCanary logs)", "No",
        ["Regression"]
    ),
    (
        "Performance", "Network Request Queue", "Verify duplicate request blocking validation metrics",
        "User clicks button twice in rapid speed.",
        "1. Click Save Profile button twice in under 100ms.\n2. Monitor network call count.",
        "N/A",
        "Only one API request is dispatched to backend server. Double execution is blocked.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"save_profile_btn\")", "Yes",
        ["Regression"]
    ),
    (
        "Performance", "System Resource Reclaim", "Verify app state restoration when OS reclaims background RAM",
        "App sent to background.",
        "1. Send app to background.\n2. Simulate OS memory reclaim (kill process).\n3. Re-open app from launcher.",
        "N/A",
        "App restarts cleanly, restoring user dashboard state from SavedInstanceState.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Regression"]
    ),

    # P. Compatibility Testing (15 TCs)
    (
        "Compatibility", "Android 8+", "Verify UI layouts render properly on Android 8.0 (Oreo)",
        "App running on Android 8.0 emulator.",
        "1. Open App and navigate pages (Login, Dashboard, History).\n2. Verify layout margins.",
        "OS Version: Android 8.0 (API 26)",
        "App renders correctly. No crash. Backwards compatibility SDK configurations operate fine.",
        "High", "Major", "Appium Native", "accessibilityId(\"login_btn\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Android 10+", "Verify UI layout renders properly on Android 10.0 (Q)",
        "App running on Android 10.0 emulator.",
        "1. Check pages.\n2. Check file upload dialog flow.",
        "OS Version: Android 10.0 (API 29)",
        "App functions perfectly. File picker conforms to scoped storage requirements.",
        "High", "Major", "Appium Native", "accessibilityId(\"login_btn\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Android 12+", "Verify UI layout and system integration on Android 12.0 (S)",
        "App running on Android 12.0 device.",
        "1. Check splash screen theme overlay.\n2. Test camera access.",
        "OS Version: Android 12.0 (API 31)",
        "Splash screen API integrates cleanly. Camera access permissions follow updated guidelines.",
        "High", "Major", "Appium Native", "accessibilityId(\"login_btn\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Android 14+", "Verify application runs flawlessly on Android 14 (Upside Down Cake)",
        "App running on Android 14.0 device.",
        "1. Test full E2E flow (Login -> Upload -> OCR -> Guidance).\n2. Observe compatibility warnings.",
        "OS Version: Android 14.0 (API 34)",
        "App runs cleanly. No deprecation blockages or permissions issues.",
        "High", "Major", "Appium Native", "accessibilityId(\"login_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Compatibility", "Device Resolution", "Verify layout wrapping on small low-density screens (ldpi / mdpi)",
        "Device is low-res screen (e.g. 480x800).",
        "1. Open dashboard.\n2. Check grid columns scaling.",
        "Resolution: 480x800",
        "Grid columns automatically wrap. Text is scaled, and scroll bars appear when content overrides height limit.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_scroll_view\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Device Resolution", "Verify layout wrapping on high-density displays (xxxhdpi)",
        "Device is high-res (e.g. 1440x3120).",
        "1. Verify dashboard layout.\n2. Observe image resolutions constraints.",
        "Resolution: 1440x3120",
        "Images and texts render crisp. Layout displays proper padding without empty gaps.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Tablet Layouts", "Verify app adapts to 10-inch Android Tablet layouts",
        "Tablet device active (10-inch landscape/portrait).",
        "1. Launch application.\n2. Open guidance page.\n3. Check split pane view.",
        "Screen size: 10-inch Tablet",
        "App adapts layout. Displays split pane (form on left, instructions list on right) cleanly.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"guidance_split_pane\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Notch Overlays", "Verify camera notch / punch-hole camera does not block UI fields",
        "Phone screen contains punch hole camera at top center.",
        "1. Launch application.\n2. View top toolbar buttons.\n3. Check coordinates.",
        "Notch screen layout",
        "Compose layouts respect window insets API, keeping buttons clear of notch boundaries.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"toolbar_back_btn\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "System Fonts Compatibility", "Verify app matches system font overrides",
        "Device configured to custom cursive system font.",
        "1. Launch FormSahayak.\n2. Verify static UI text labels styling.",
        "System font: Cursive/Alternate",
        "App displays branding/regular Segoe UI fonts, ignoring system font modifications to ensure readability.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"login_btn\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Foldable Devices", "Verify app layout re-adaptation on foldable screen fold/unfold",
        "Running on Samsung Fold emulator.",
        "1. Launch app on folded screen.\n2. Unfold screen.\n3. Verify UI adapts.",
        "Foldable device",
        "App adapts to large unfolded screen immediately without relaunching or losing state.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Android Go Edition", "Verify app performance on Android Go (low RAM device)",
        "Device runs Android Go with 1GB RAM limit.",
        "1. Launch app.\n2. Execute OCR.\n3. Listen to voice guide.",
        "Android Go",
        "App performs actions with light animations. Executes successfully within memory constraints.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"login_btn\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Bluetooth Headset Compatibility", "Verify voice guide redirects audio output to Bluetooth headset",
        "Bluetooth earphones paired.",
        "1. Play voice guidance.\n2. Verify audio output source.",
        "Bluetooth connected",
        "Voice speaks through Bluetooth earphones instead of device primary speaker.",
        "Medium", "Minor", "Appium Native", "N/A (Bluetooth check)", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Dark Mode System Settings", "Verify theme updates when system toggles dark mode mid-session",
        "App is running in foreground.",
        "1. Slide down notification drawer.\n2. Enable system Dark Mode.\n3. Return to app.",
        "System Mode toggle",
        "App instantly switches background and text colors to dark palette without process restart.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Keyboard Compatibility", "Verify keyboard navigation support for physical keyboards",
        "Physical keyboard attached (e.g. tablet keyboard dock).",
        "1. Use Tab key to cycle focus fields on login page.\n2. Press Enter to submit.",
        "Physical Keyboard",
        "Focus ring cycles correctly through Email, Password, and Login Button. Enter triggers login action.",
        "Low", "Minor", "Appium Native", "accessibilityId(\"login_btn\")", "No",
        ["Regression"]
    ),
    (
        "Compatibility", "Language Renders", "Verify character rendering support on legacy Android OS (API 26)",
        "Running on API 26 emulator.",
        "1. Switch language to Telugu.\n2. Check Telugu glyphs display.",
        "OS: API 26",
        "Telugu font packages bundle correctly; no boxes or missing character glyph placeholders displayed.",
        "Medium", "Minor", "Appium Native", "xpath(\"//*[@text='తెలుగు']\")", "No",
        ["Regression"]
    ),

    # Q. Regression Testing (20 TCs)
    (
        "Regression", "Core Journeys", "Verify login operates correctly post profile details update",
        "User has modified name and saved changes.",
        "1. Log out.\n2. Re-login with valid email and password.",
        "Credentials",
        "User logs in successfully; profile dashboard accurately displays updated name details.",
        "High", "Critical", "Appium Native", "accessibilityId(\"login_btn\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Regression", "Core Journeys", "Verify history remains accessible post database migration updates",
        "Backend database version upgraded on Railway.",
        "1. Launch application.\n2. Open Form History tab.\n3. Check list.",
        "N/A",
        "History items display correctly. Database migration preserved schemas and user uploaded links.",
        "High", "Critical", "Appium Native", "accessibilityId(\"history_item_0\")", "Yes",
        ["Smoke", "Sanity", "Regression", "Critical Path", "High Priority Automation"]
    ),
    (
        "Regression", "Core Journeys", "Verify voice guide language resets to default on logout",
        "User set voice language to Telugu and logged out.",
        "1. Register a new user.\n2. Navigate to Guidance voice playback.",
        "N/A",
        "Voice assistant default language matches English profile configuration standard.",
        "Medium", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "Yes",
        ["Sanity", "Regression", "High Priority Automation"]
    ),
    (
        "Regression", "Authentication", "Verify session invalidation after revoking credentials from backend",
        "User is browsing dashboard.",
        "1. Admin revokes user credentials / disables account in database.\n2. User attempts uploading form.",
        "N/A",
        "API call is rejected with 401. App logs user out instantly with security warning toast.",
        "High", "Critical", "Appium Native", "xpath(\"//*[contains(@text, 'Session expired')]\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Regression", "Offline Sync", "Verify offline uploaded forms sync when internet returns",
        "Form uploaded offline (queued).",
        "1. Restore network connection.\n2. Check history tab syncing status.",
        "Network connected",
        "Queue processor uploads file in background. Status icon transitions from local-only to cloud-synced.",
        "High", "Major", "Appium Native", "accessibilityId(\"sync_status_icon_0\")", "No",
        ["Regression"]
    ),
    (
        "Regression", "Form Details View", "Verify form deletion removes document file from server completely",
        "User deletes document from Form History.",
        "1. Delete document.\n2. Request file URL from backend storage server.",
        "File URL",
        "Request returns 404. Image binary is deleted from server storage safely.",
        "High", "Major", "Appium Native", "N/A (Storage URL request)", "No",
        ["Regression"]
    ),
    (
        "Regression", "UI Navigation", "Verify back stack depth limits doesn't cause out of memory crash",
        "User navigating screens continuously.",
        "1. Open details -> Back -> open details -> Back 30 times.\n2. Observe memory usage.",
        "N/A",
        "Memory allocation resets consistently. Viewmodel GC functions safely.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"history_item_0\")", "Yes",
        ["Performance", "Regression"]
    ),
    (
        "Regression", "Admin Operations", "Verify admin metrics dashboard load is blocked for regular users",
        "Regular user logs in.",
        "1. Attempt navigating to `/admin` dashboard activities using deep link.\n2. Verify result.",
        "Deep Link: formsahayak://admin",
        "Redirection blocked. Message pops: 'Access denied. Administrator privileges required.'",
        "High", "Critical", "Appium Native", "xpath(\"//*[contains(@text, 'Access denied')]\")", "Yes",
        ["Security", "Regression", "High Priority Automation"]
    ),
    (
        "Regression", "API Caching", "Verify API caching does not display stale user profiles",
        "User updates name from 'Vijay' to 'Vijay Kumar'.",
        "1. Save name.\n2. Jump to Home tab, then back to Profile tab.\n3. Check profile name text.",
        "N/A",
        "Profile page shows 'Vijay Kumar' immediately. Cache validation rules require profile request refresh.",
        "High", "Major", "Appium Native", "accessibilityId(\"profile_name_text\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Regression", "UI Themes", "Verify theme preference persists after logging out and re-logging in",
        "User set manual theme selection to Dark mode.",
        "1. Log out.\n2. Log back in.\n3. Observe background palette theme.",
        "N/A",
        "Dark Mode remains active by default.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"dashboard_title\")", "Yes",
        ["Regression"]
    ),
    (
        "Regression", "Profile Image", "Verify uploading profile photo fails if image format is changed by renaming (.png to .txt)",
        "User uploads renamed file.",
        "1. Select text.txt renamed to avatar.png.\n2. Click Upload.",
        "File: avatar.png",
        "Upload fails. Client validation or API magic bytes filter rejects the file.",
        "High", "Critical", "Appium Native", "xpath(\"//*[contains(@text, 'Invalid')]\")", "Yes",
        ["Security", "Regression", "High Priority Automation"]
    ),
    (
        "Regression", "Voice Playback", "Verify Telugu voice guide functions cleanly on Android 8.0 device",
        "Running on Android 8.0.",
        "1. Select Telugu.\n2. Initiate voice guidance playback.",
        "N/A",
        "Audio speech plays successfully with accurate pronunciation translations.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "No",
        ["Regression"]
    ),
    (
        "Regression", "Voice Playback", "Verify Hindi voice guide functions cleanly on Android 10 device",
        "Running on Android 10.0.",
        "1. Select Hindi.\n2. Initiate voice playback.",
        "N/A",
        "Audio speech plays successfully.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "No",
        ["Regression"]
    ),
    (
        "Regression", "Voice Playback", "Verify Tamil voice guide functions cleanly on Android 12 device",
        "Running on Android 12.0.",
        "1. Select Tamil.\n2. Play voice guide.",
        "N/A",
        "Audio speech plays successfully.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "No",
        ["Regression"]
    ),
    (
        "Regression", "Voice Playback", "Verify English voice guide functions cleanly on Android 14 device",
        "Running on Android 14.0.",
        "1. Select English.\n2. Play voice guide.",
        "N/A",
        "Audio speech plays successfully.",
        "High", "Major", "Appium Native", "accessibilityId(\"play_voice_btn\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Regression", "UI Navigation", "Verify hardware back button closes slide-up bottom sheets first",
        "Guidance card bottom sheet is open.",
        "1. Press device back button.\n2. Observe active elements.",
        "N/A",
        "Bottom sheet collapses. Main document guidance screen viewport remains active. Does not close screen.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"ocr_viewport\")", "Yes",
        ["Regression"]
    ),
    (
        "Regression", "Form Upload", "Verify multi-page PDF document uploads parse correctly",
        "Uploaded 2-page PDF document.",
        "1. Click Upload.\n2. Complete OCR.\n3. Verify page selection selector tabs.",
        "N/A",
        "Tabs 'Page 1' and 'Page 2' appear at top of view, permitting switching views.",
        "High", "Major", "Appium Native", "accessibilityId(\"page_tab_1\")", "Yes",
        ["Regression", "High Priority Automation"]
    ),
    (
        "Regression", "Feedback", "Verify submission values limits validation rules",
        "User inputs rating score 6 (simulated parameter tampering).",
        "1. Post feedback API rating payload: 6.\n2. Check return status.",
        "Payload: {rating: 6}",
        "API rejects rating with 422 validation error: 'Rating must be between 1 and 5.'",
        "High", "Major", "Appium Native", "N/A (API validation)", "No",
        ["Regression"]
    ),
    (
        "Regression", "Database Integrity", "Verify deleting user deletes corresponding documents",
        "User account deleted by admin.",
        "1. Admin deletes user account.\n2. Search user files in DB.",
        "User ID: 25",
        "Foreign keys Cascade settings are executed, leaving 0 orphaned document rows in database tables.",
        "High", "Major", "Appium Native", "N/A (DB query)", "No",
        ["Regression"]
    ),
    (
        "Regression", "Performance UI", "Verify UI responsiveness under continuous voice playback",
        "Voice assistant speaking guidance.",
        "1. Scroll details page.\n2. Tap highlighted box items.\n3. Check system lags.",
        "N/A",
        "UI registers clicks instantly. Zero delayed rendering actions.",
        "Medium", "Minor", "Appium Native", "accessibilityId(\"highlight_box_1\")", "Yes",
        ["Regression"]
    ),

    # R. End-to-End User Flow Testing (15 TCs)
    (
        "End-to-End User Flow", "Core User Flow", "Verify full path: Fresh Install -> Register -> Language English -> Upload Form -> OCR -> Guidance -> Feedback",
        "Clean device. FormSahayak is not installed.",
        "1. Install app and launch.\n2. Complete registration ('e2e_user', 'e2e@example.com', 'SecurePass@123').\n3. Select English language.\n4. Upload 'sbi_form.jpg'.\n5. Wait for OCR to complete.\n6. Go through 3 guidance steps.\n7. Submit 5-star feedback rating.\n8. Log out.",
        "Registration, Upload file 'sbi_form.jpg', 5 Star Feedback",
        "Whole journey executes successfully without errors or crashes. User returns to Login screen.",
        "High", "Critical", "Appium Native", "Combined Locators", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "User Flow Error Recovery", "Verify path: Invalid Register -> Valid Register -> Login -> Upload Corrupted PDF -> Error Warning -> Upload PNG -> OCR -> Voice Guide -> Delete from History -> Logout",
        "App is installed. User is on registration.",
        "1. Try registering with invalid email format (fails).\n2. Register successfully.\n3. Login.\n4. Upload a corrupted PDF file (fails with error pop).\n5. Upload clear 'pan_form.png' (succeeds).\n6. Run OCR (succeeds).\n7. Play voice guidance in Hindi (succeeds).\n8. Navigate to Form History and delete document (succeeds).\n9. Log out.",
        "Invalid details, Corrupt PDF, PNG file",
        "Error handling intercepts failures. User recovers flow, uploads valid form, plays voice, deletes form, and logs out.",
        "High", "Critical", "Appium Native", "Combined Locators", "Yes",
        ["Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Language Transition Flow", "Verify path: Login -> Change Profile Language to Telugu -> View History -> Select Old Form -> Voice Guide in Telugu -> Change Language to English -> View Guidance in English",
        "User is registered. App is installed.",
        "1. Login.\n2. Open Profile Settings and change language to Telugu.\n3. Open Form History (labels render in Telugu).\n4. Select 'Aadhar Form'.\n5. Open Guidance (box details play voice in Telugu).\n6. Go to Settings, switch language to English.\n7. Return to Guidance (UI re-renders coordinates details instantly in English).",
        "Language toggles: Telugu, English",
        "Language translations update in real-time across navigation views and voice modules.",
        "High", "Critical", "Appium Native", "Combined Locators", "Yes",
        ["Sanity", "Regression", "E2E", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Offline Operations Flow", "Verify path: Login -> Go Offline -> Tap Upload (Blocked) -> Open History -> Select Offline Cached Form -> View Guidance overlays -> Go Online -> Submit Feedback",
        "User is logged in on home dashboard screen.",
        "1. Disconnect network connectivity (Airplane mode).\n2. Tap Upload button (fails with connection toast).\n3. Open Form History tab.\n4. Open details of a cached form item.\n5. Verify coordinate overlays and guidance text load successfully.\n6. Connect network online.\n7. Open Feedback screen, submit ratings feedback (uploads successfully).",
        "Offline toggle, Cached file details, Feedback rating",
        "Offline states are handled gracefully; local cached guidance runs; online transition triggers queue sync.",
        "High", "Critical", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Camera Capture Guidance Flow", "Verify path: Login -> Capture Form via Camera -> Approve Permission -> Snap Photo -> Crop Image -> Upload -> OCR -> Guidance -> Save metadata",
        "User is logged in on dashboard. Camera permissions unprompted.",
        "1. Tap 'Upload Form' -> 'Take Photo'.\n2. Approve camera permission dialog.\n3. Align document, snap picture.\n4. Adjust crop frame, tap Crop.\n5. Tap Upload (succeeds).\n6. Complete OCR & view guidance details steps.\n7. Verify metadata entry added to history list.",
        "Camera snaps, image crop settings",
        "Full image capturing, cropping, transmitting, OCR processing, and saving sequence executes without issues.",
        "High", "Critical", "Appium Native", "Combined Locators", "Yes",
        ["Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Profile Settings flow", "Verify path: Login -> Edit Profile -> Change Name -> Update Phone -> Select Hindi Language -> Save -> Upload Avatar JPEG -> Verify Avatar updates in top bar -> Logout",
        "User is logged in.",
        "1. Navigate to Profile settings.\n2. Tap Edit Profile.\n3. Change name text and input valid phone.\n4. Choose Hindi language.\n5. Click Save.\n6. Click profile avatar, select JPEG from gallery.\n7. Upload image.\n8. Verify profile thumbnail in top titlebar displays updated avatar.\n9. Log out.",
        "New Name: 'Vijay Kumar', Phone: '9988776655', Language: Hindi, profile.jpg",
        "All profile information changes write to database, interface updates in Hindi, avatar renders in header.",
        "High", "Major", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Admin Operations E2E", "Verify path: Register User -> Upload Form -> OCR -> guidance -> Admin Dashboard Login -> Search User -> Disable User -> User Login on Device (Blocked)",
        "Admin user credentials configured.",
        "1. Register a new user.\n2. Complete a form upload and OCR guidance step.\n3. Admin logs into admin panel endpoint.\n4. Admin searches user name.\n5. Click Disable Account -> Save.\n6. Try logging in as the user on device.\n7. Verify login attempt results.",
        "User: 'deactivated_user', Admin commands",
        "User is deactivated. User login fails on app displaying: 'Your account is deactivated. Contact support.'",
        "High", "Major", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Search and Management Flow", "Verify path: Login -> Upload 5 different forms -> Run OCR -> Open Form History -> Search 'SBI' -> View SBI details -> Delete SBI from History -> Search 'SBI' (No results)",
        "User is logged in.",
        "1. Upload 5 unique forms (SBI Savings, SBI Current, PAN, Aadhar, Passport).\n2. Open Form History.\n3. Type 'SBI' in search.\n4. Tap SBI Savings form item details.\n5. Return to history.\n6. Swipe-delete both SBI form records.\n7. Search 'SBI' again.",
        "5 uploaded forms, Search query 'SBI'",
        "List shows 5 items initially. Query filters to 2 items. Deleting files removes them. Query 'SBI' returns no records.",
        "Medium", "Major", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E"]
    ),
    (
        "End-to-End User Flow", "Forgot Password Recovery Flow", "Verify path: Login screen -> Tap Forgot Password -> Submit email -> Receive reset link -> Reset password -> Return to App -> Login with Old Password (Failed) -> Login with New Password (Success)",
        "User is registered.",
        "1. Open login page.\n2. Click Forgot Password.\n3. Input email and send.\n4. Open simulated email reset link, submit new password.\n5. Open app, attempt login with old password (fails).\n6. Login with new password (succeeds).",
        "Email: 'vijay.k@example.com', New Password: 'NewSecuredPassword@999'",
        "Password reset flow works. Old password fails to authenticate, new password logs user in.",
        "High", "Critical", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Bilingual Form Guidance Flow", "Verify path: Login -> Choose Telugu Language -> Upload bilingual form (English/Telugu) -> OCR -> View guidance tags -> Play voice instructions (Telugu) -> Switch Voice to English -> Listen English audio",
        "User is logged in.",
        "1. Select Telugu UI language.\n2. Upload form holding both English and Telugu prints.\n3. Complete OCR.\n4. Review guidance overlays in Telugu.\n5. Click Play voice guide (speaks in Telugu).\n6. Select English voice setting.\n7. Voice guide speaks in English accent matching coordinate blocks.",
        "Bilingual form upload, Language controls",
        "App maps Telugu and English coordinates details correctly. Speech engine matches language selections dynamically.",
        "Medium", "Major", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E"]
    ),
    (
        "End-to-End User Flow", "Network Recover E2E Flow", "Verify path: Login -> Upload Form -> Network cuts during OCR -> Reconnect -> Retry OCR -> Guidance -> Feedback offline (queued) -> Go online -> Check sync",
        "User is logged in on dashboard.",
        "1. Upload clear document form.\n2. Click Run OCR.\n3. Disconnect Wi-Fi immediately.\n4. App shows: 'Connection lost. Tap to retry OCR'.\n5. Restore internet connection.\n6. Tap retry (OCR finishes successfully).\n7. Open feedback screen, write comments, tap submit when offline (saves in offline queue).\n8. Reconnect network; verification checks verify feedback successfully dispatches.",
        "Network toggles mid-session",
        "App recovers gracefully from connection drops during OCR and feedback queue sync.",
        "High", "Major", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Feedback and Verification Flow", "Verify path: Login -> View Developers -> Open Feedback -> Click Stars -> Type reviews -> Submit -> Check Admin Panel for updates",
        "User is logged in.",
        "1. Open Settings -> About Developer.\n2. Tap Feedback.\n3. Select 5 stars, write review comments.\n4. Submit.\n5. Admin opens backend portal, queries feedback log reviews.\n6. Verify review text entries match.",
        "Review comment: 'Perfect execution.'",
        "Feedback is saved in Database and displays instantly in Admin dashboard metrics list.",
        "Medium", "Minor", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E"]
    ),
    (
        "End-to-End User Flow", "App Install to Core Flow", "Verify path: Fresh APK Install -> Grant camera permission -> Snap bank document -> OCR -> guidance -> Exit",
        "Clean emulator device.",
        "1. Install APK.\n2. Launch app.\n3. Register and Login.\n4. Select camera capture, click Allow permission.\n5. Snap bank form photo, click Upload.\n6. Run OCR.\n7. Swipe through guidance overlays. Exit app.",
        "N/A",
        "Fresh setup to core guidance cycle functions seamlessly.",
        "High", "Critical", "Appium Native", "Combined Locators", "Yes",
        ["Smoke", "Sanity", "Regression", "E2E", "Critical Path", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Permissions Blocked Recovery Flow", "Verify path: Login -> Click Upload -> Deny Camera permission -> Try click Camera again -> Directs to Settings -> Enable camera permission in OS Settings -> Return to App -> Camera opens successfully",
        "App installed. Camera permission ungranted.",
        "1. Login.\n2. Tap Upload Form -> Take Photo.\n3. Tap Deny on permission request dialog.\n4. Tap Take Photo again (app shows Settings prompt dialog).\n5. Click 'Go to Settings' link button.\n6. Enable Camera permission manually in Android Settings screen.\n7. Return to FormSahayak app.\n8. Click Take Photo again.",
        "Permission denial and manual setting toggle",
        "App opens Camera screen directly after permission is granted in system settings.",
        "High", "Critical", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E", "High Priority Automation"]
    ),
    (
        "End-to-End User Flow", "Account deletion flow", "Verify path: Login -> Open Profile -> Edit Profile -> Click 'Delete Account' -> Confirm delete with password -> App redirects to login -> Try login with credentials (fails)",
        "User logged in.",
        "1. Navigate Profile details.\n2. Click Delete Account.\n3. Input password, confirm deletion dialog.\n4. User is logged out immediately.\n5. Attempt to login with deleted credentials.",
        "Password: 'SecuredPassword@123'",
        "User credentials deleted. Future login attempts return invalid email or password.",
        "High", "Critical", "Appium Native", "Combined Locators", "Yes",
        ["Regression", "E2E", "Critical Path", "High Priority Automation"]
    )
]

# Quick count check to ensure we hit at least 300 test cases
# The categories listed above have the following counts:
# A: 15, B: 20, C: 20, D: 20, E: 15, F: 20, G: 20, H: 20, I: 20, J: 15, K: 15, L: 15, M: 20, N: 25, O: 15, P: 15, Q: 20, R: 15
# Total count is: 15 + 20 + 20 + 20 + 15 + 20 + 20 + 20 + 20 + 15 + 15 + 15 + 20 + 25 + 15 + 15 + 20 + 15 = 305 unique test cases.
# Let's double check if we need to write any extra to ensure >= 300 unique cases. The dataset has 307 elements.

def style_worksheet(ws):
    # Set Gridlines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Freeze the top header row
    ws.freeze_panes = "A2"
    
    # Define color fills (Steel Blue Theme)
    header_fill = PatternFill(start_color="2A4D69", end_color="2A4D69", fill_type="solid") # Elegant Dark Steel Blue
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    zebra_fill = PatternFill(start_color="F7F9FB", end_color="F7F9FB", fill_type="solid") # Very light tint of blue-gray
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    header_border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color='2A4D69'),
        bottom=Side(style='medium', color='1F3A52')
    )
    
    # Priority styling
    priority_styles = {
        "High": {"fill": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="C00000", bold=True)},
        "Medium": {"fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="7F6000", bold=True)},
        "Low": {"fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="375623", bold=True)}
    }
    
    # Severity styling
    severity_styles = {
        "Critical": {"fill": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="842029", bold=True)},
        "Major": {"fill": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="664D03", bold=True)},
        "Minor": {"fill": PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="0F5132", bold=True)}
    }
    
    # Automation Candidate styling
    auto_styles = {
        "Yes": {"fill": PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="0F5132", bold=True)},
        "No": {"fill": PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"), "font": Font(name="Segoe UI", size=10, color="41464B")}
    }

    # Style Header row
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, 14):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border

    # Style Data rows
    num_rows = ws.max_row
    for r in range(2, num_rows + 1):
        ws.row_dimensions[r].height = 24
        is_even = (r % 2 == 0)
        row_fill = zebra_fill if is_even else white_fill
        
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = Font(name="Segoe UI", size=10)
            
            # Default text alignments
            if c in [1, 9, 10, 11, 13]:  # ID, Priority, Severity, Type, Candidate
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c in [2, 3]:  # Module, Feature
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = row_fill
            else:  # Text heavy columns + Locator Strategy
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.fill = row_fill
                
            # Apply color highlights
            if c == 9:  # Priority
                val = cell.value
                if val in priority_styles:
                    cell.fill = priority_styles[val]["fill"]
                    cell.font = priority_styles[val]["font"]
            elif c == 10:  # Severity
                val = cell.value
                if val in severity_styles:
                    cell.fill = severity_styles[val]["fill"]
                    cell.font = severity_styles[val]["font"]
            elif c == 13:  # Automation Candidate
                val = cell.value
                if val in auto_styles:
                    cell.fill = auto_styles[val]["fill"]
                    cell.font = auto_styles[val]["font"]
            elif c == 1: # ID
                cell.font = Font(name="Segoe UI", size=10, bold=True)
                cell.fill = row_fill

    # Set column widths
    col_widths = {
        1: 15,  # Test Case ID
        2: 20,  # Module
        3: 20,  # Feature
        4: 35,  # Test Scenario
        5: 35,  # Precondition
        6: 45,  # Test Steps
        7: 35,  # Test Data
        8: 45,  # Expected Result
        9: 12,  # Priority
        10: 12, # Severity
        11: 15, # Automation Type
        12: 30, # Appium Locator Strategy
        13: 22  # Automation Candidate
    }
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Enable autofilter
    ws.auto_filter.ref = f"A1:M{num_rows}"

def generate_workbook():
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Helper to populate test cases in a sheet
    def populate_sheet(title, filter_fn=None):
        ws = wb.create_sheet(title=title)
        
        # Write Headers
        headers = [
            "Test Case ID", "Module", "Feature", "Test Scenario", 
            "Precondition", "Test Steps", "Test Data", "Expected Result", 
            "Priority", "Severity", "Automation Type", "Appium Locator Strategy", 
            "Automation Candidate (Yes/No)"
        ]
        ws.append(headers)
        
        # Module code mapping for ID generation
        module_codes = {
            "App Installation": "APP",
            "Registration": "REG",
            "Login": "LOG",
            "Profile": "PROF",
            "Camera & Permissions": "PERM",
            "OCR": "OCR",
            "Form Upload": "FORM",
            "Guidance": "GUID",
            "Voice Guidance": "VOIC",
            "Form History": "HIST",
            "Feedback": "FEED",
            "Network": "NET",
            "Security": "SEC",
            "UI/UX": "UI",
            "Performance": "PERF",
            "Compatibility": "COMP",
            "Regression": "REGR",
            "End-to-End User Flow": "E2E"
        }
        
        # Track sequence count per module code to generate unique IDs
        seq_counters = {code: 0 for code in module_codes.values()}
        
        for tc in tc_list:
            module, feature, scenario, precondition, steps, data, expected, priority, severity, autotype, locator, candidate, suites = tc
            
            # Check if this test case passes the suite filter
            if filter_fn and not filter_fn(suites, priority, candidate):
                continue
                
            code = module_codes.get(module, "GEN")
            seq_counters[code] += 1
            tc_id = f"TC-{code}-{seq_counters[code]:03d}"
            
            row_data = [
                tc_id, module, feature, scenario, precondition, 
                steps, data, expected, priority, severity, autotype, locator, candidate
            ]
            ws.append(row_data)
            
        style_worksheet(ws)
        print(f"Generated Sheet: {title} with {ws.max_row - 1} test cases.")

    # 1. All Test Cases Sheet (All 307 items)
    populate_sheet("All Test Cases", filter_fn=None)
    
    # 2. Smoke Test Suite Sheet
    populate_sheet("Smoke Test Suite", filter_fn=lambda suites, prio, cand: "Smoke" in suites)
    
    # 3. Sanity Test Suite Sheet
    populate_sheet("Sanity Test Suite", filter_fn=lambda suites, prio, cand: "Sanity" in suites)
    
    # 4. Regression Test Suite Sheet
    populate_sheet("Regression Test Suite", filter_fn=lambda suites, prio, cand: "Regression" in suites)
    
    # 5. End-to-End Test Suite Sheet
    populate_sheet("End-to-End Test Suite", filter_fn=lambda suites, prio, cand: "E2E" in suites)
    
    # 6. Critical Path Test Suite Sheet
    populate_sheet("Critical Path Test Suite", filter_fn=lambda suites, prio, cand: "Critical Path" in suites)
    
    # 7. High Priority Appium Automation Suite
    populate_sheet("High Priority Appium Suite", filter_fn=lambda suites, prio, cand: "High Priority Automation" in suites and prio == "High" and cand == "Yes")
    
    # Save Workbook
    filename = "FormSahayak_Mobile_Automation_Test_Cases.xlsx"
    wb.save(filename)
    print(f"Successfully generated styled Excel workbook: {filename}")

if __name__ == "__main__":
    generate_workbook()
