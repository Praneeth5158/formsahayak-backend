import os
import sys
import shutil
import io
import time
from unittest.mock import MagicMock
from datetime import datetime

# 1. Force testing mode using environment variable before any imports
os.environ["TESTING"] = "True"

# Setup dummy environment variables for testing if not already present
os.environ.setdefault("ADMIN_EMAIL", "admin@formsahayak.com")
os.environ.setdefault("ADMIN_PASSWORD", "AdminPassword123!")
os.environ.setdefault("JWT_SECRET", "super_secret_formsahayak_admin_key_2026")
os.environ.setdefault("GROQ_API_KEY", "dummy_groq_key_for_testing")

# Clean up previous test database
TEST_DB_PATH = "test.db"
if os.path.exists(TEST_DB_PATH):
    try:
        os.remove(TEST_DB_PATH)
    except Exception as e:
        print(f"Could not remove old test database: {e}")

# 2. Mock external dependencies at module level before importing main
import pytesseract
pytesseract.image_to_data = MagicMock(return_value={
    "text": ["Name", "Phone", "Email", "Signature"],
    "left": [10, 20, 30, 40],
    "top": [10, 20, 30, 40],
    "width": [50, 50, 50, 50],
    "height": [15, 15, 15, 15],
    "conf": [90, 95, 85, 99]
})

import main
from main import app
from db import SessionLocal, engine
from models import Base, User, Document, Feedback, DeveloperDetails

# Create SQLite tables
Base.metadata.create_all(bind=engine)

# Mock Groq client
main.client = MagicMock()
mock_groq_response = MagicMock()
mock_groq_response.choices = [MagicMock()]
mock_groq_response.choices[0].message.content = "Step 1. Fill your Name (Name) in the first box. Step 2. Sign at the bottom (Signature)."
main.client.chat.completions.create.return_value = mock_groq_response

# Mock send_otp_email_helper to prevent API calls to Resend/Brevo
main.send_otp_email_helper = MagicMock()

# Mock async background task for media generation to run synchronously/do nothing
async def dummy_generate_media_background(*args, **kwargs):
    pass
main.generate_media_background = dummy_generate_media_background

# Mock SMTP email sending if fallback smtplib is ever used
import smtplib
smtplib.SMTP = MagicMock()

# 3. Create the FastAPI TestClient
from fastapi.testclient import TestClient
client = TestClient(app)

# Track created test files for cleanup later
created_files = []

# List to collect all test results
test_results = []

def record_result(endpoint, method, test_case, process, response, custom_status=None, custom_content=None):
    """Helper to record test outcome."""
    status = "PASS"
    code = response.status_code if response is not None else 0
    content = ""
    
    if response is not None:
        try:
            content = response.text[:250] + ("..." if len(response.text) > 250 else "")
        except Exception:
            content = "[Binary/Unreadable]"
    
    if custom_status:
        status = custom_status
    elif response is not None:
        # Check standard success codes
        if response.status_code >= 400:
            # Some test cases test validation errors (expected 400 or 404)
            # If the test case is testing validation, status is PASS. Otherwise FAIL.
            is_error_test = any(x in test_case.lower() for x in ["fail", "error", "mismatch", "weak", "invalid", "unregistered", "not found", "unauthorized"])
            if is_error_test:
                status = "PASS"
            else:
                status = "FAIL"
        else:
            # 2xx / 3xx response.
            is_success_test = not any(x in test_case.lower() for x in ["fail", "error", "mismatch", "weak", "invalid", "unregistered", "not found", "unauthorized"])
            if is_success_test:
                status = "PASS"
            else:
                # Expected error but got success
                status = "FAIL"
    else:
        status = "FAIL"

    if custom_content:
        content = custom_content

    test_results.append({
        "Endpoint": endpoint,
        "Method": method,
        "Test Case": test_case,
        "Testing Process": process,
        "Status": status,
        "Response Code": code,
        "Response Content": content
    })
    
    print(f"[{status}] {method} {endpoint} - {test_case} (Code: {code})")
    return status

# ==============================================================================
# TEST SUITE RUN
# ==============================================================================
print("\n--- STARTING API ENDPOINTS TESTING ---\n")

# --- 1. SIGNUP ENDPOINT TESTS ---
# Case 1.1: Successful signup
process = "Send POST request to /signup with a new email, name, strong password, and matching confirm_password."
payload = {
    "name": "Test User",
    "email": "testuser@example.com",
    "password": "Password123!",
    "confirm_password": "Password123!"
}
res = client.post("/signup", json=payload)
record_result("/signup", "POST", "Signup - Successful registration", process, res)

# Case 1.2: Mismatched password
process = "Send POST request to /signup with passwords that do not match."
payload["confirm_password"] = "Mismatch123!"
res = client.post("/signup", json=payload)
record_result("/signup", "POST", "Signup - Password mismatch error", process, res)

# Case 1.3: Weak password
process = "Send POST request to /signup with password that violates validation rules (no uppercase/number/special char)."
payload["password"] = "weak"
payload["confirm_password"] = "weak"
res = client.post("/signup", json=payload)
record_result("/signup", "POST", "Signup - Weak password validation error", process, res)

# Case 1.4: Duplicate email
process = "Send POST request to /signup with an email address that is already registered."
payload = {
    "name": "Another Name",
    "email": "testuser@example.com",
    "password": "Password123!",
    "confirm_password": "Password123!"
}
res = client.post("/signup", json=payload)
record_result("/signup", "POST", "Signup - Duplicate email registration error", process, res)


# --- 2. LOGIN ENDPOINT TESTS ---
# Case 2.1: Login success
process = "Send POST request to /login with correct email and password."
payload = {
    "email": "testuser@example.com",
    "password": "Password123!"
}
res = client.post("/login", json=payload)
record_result("/login", "POST", "Login - Success with valid credentials", process, res)

# Case 2.2: Invalid email
process = "Send POST request to /login with an unregistered email."
payload["email"] = "fake@example.com"
res = client.post("/login", json=payload)
record_result("/login", "POST", "Login - Fail on unregistered email", process, res)

# Case 2.3: Invalid password
process = "Send POST request to /login with correct email but wrong password."
payload = {
    "email": "testuser@example.com",
    "password": "WrongPassword123"
}
res = client.post("/login", json=payload)
record_result("/login", "POST", "Login - Fail on incorrect password", process, res)


# --- 3. PROFILE ENDPOINT TESTS ---
# Case 3.1: Get profile success
process = "Send GET request to /profile/{email} for a registered user."
res = client.get("/profile/testuser@example.com")
record_result("/profile/{email}", "GET", "Get Profile - Success for existing user", process, res)

# Case 3.2: Get profile user not found
process = "Send GET request to /profile/{email} for a non-registered user."
res = client.get("/profile/nonexistent@example.com")
record_result("/profile/{email}", "GET", "Get Profile - Fail for non-existent user", process, res)

# Case 3.3: Update profile success
# Need to send form data with optional files
process = "Send POST request to /update-profile with email, phone, language, and mock profile image."
from PIL import Image
import io
img = Image.new("RGB", (10, 10), color="blue")
img_bytes = io.BytesIO()
img.save(img_bytes, format="JPEG")
img_bytes.seek(0)

files = {"profile_image": ("profile.jpg", img_bytes, "image/jpeg")}
data = {
    "email": "testuser@example.com",
    "phone": "9876543210",
    "language": "Hindi"
}
res = client.post("/update-profile", data=data, files=files)
record_result("/update-profile", "POST", "Update Profile - Success updating email details", process, res)

# Case 3.4: Update profile user not found
process = "Send POST request to /update-profile for an unregistered email."
data["email"] = "fake@example.com"
res = client.post("/update-profile", data=data, files=files)
record_result("/update-profile", "POST", "Update Profile - Fail for non-existent user", process, res)


# --- 4. OTP / PASSWORD RESET TESTS ---
# Case 4.1: Forgot password success
process = "Send POST request to /forgot-password with registered email to trigger OTP."
res = client.post("/forgot-password", json={"email": "testuser@example.com"})
record_result("/forgot-password", "POST", "Forgot Password - Success triggering OTP", process, res)

# Case 4.2: Forgot password email not found
process = "Send POST request to /forgot-password with unregistered email."
res = client.post("/forgot-password", json={"email": "fake@example.com"})
record_result("/forgot-password", "POST", "Forgot Password - Fail for non-existent user", process, res)

# Case 4.3: Send OTP success
process = "Send POST request to /send-otp with registered email to trigger OTP."
res = client.post("/send-otp", json={"email": "testuser@example.com"})
record_result("/send-otp", "POST", "Send OTP - Success generating OTP", process, res)

# Retrieve the OTP that was saved in memory
stored_otp = main.otp_storage.get("testuser@example.com")

# Case 4.4: Verify OTP success
process = "Send POST request to /verify-otp with correct email and the OTP generated in-memory."
res = client.post("/verify-otp", json={"email": "testuser@example.com", "otp": stored_otp})
record_result("/verify-otp", "POST", "Verify OTP - Success with correct OTP", process, res)

# Case 4.5: Verify OTP invalid OTP
process = "Send POST request to /verify-otp with correct email but invalid OTP."
res = client.post("/verify-otp", json={"email": "testuser@example.com", "otp": "9999"})
record_result("/verify-otp", "POST", "Verify OTP - Fail with invalid OTP code", process, res)

# Case 4.6: Verify OTP not found
process = "Send POST request to /verify-otp with email that has no active OTP."
res = client.post("/verify-otp", json={"email": "fake@example.com", "otp": "1234"})
record_result("/verify-otp", "POST", "Verify OTP - Fail when OTP doesn't exist", process, res)

# Case 4.7: Reset password success
process = "Send POST request to /reset-password with verified email, new password, and matching confirm_password."
res = client.post("/reset-password", json={
    "email": "testuser@example.com",
    "new_password": "NewPassword123!",
    "confirm_password": "NewPassword123!"
})
record_result("/reset-password", "POST", "Reset Password - Success with matching password", process, res)

# Case 4.8: Reset password mismatched confirm password
process = "Send POST request to /reset-password with mismatched passwords."
res = client.post("/reset-password", json={
    "email": "testuser@example.com",
    "new_password": "NewPassword123!",
    "confirm_password": "MismatchPassword1!"
})
record_result("/reset-password", "POST", "Reset Password - Fail with mismatched confirm password", process, res)

# Case 4.9: Reset password user not found
process = "Send POST request to /reset-password with unregistered email."
res = client.post("/reset-password", json={
    "email": "fake@example.com",
    "new_password": "NewPassword123!",
    "confirm_password": "NewPassword123!"
})
record_result("/reset-password", "POST", "Reset Password - Fail for non-existent user", process, res)


# --- 5. DOCUMENT UPLOAD & DOCUMENT HISTORY TESTS ---
# Let's create a small valid PNG file
img_100 = Image.new("RGB", (100, 100), color="white")
img_100_bytes = io.BytesIO()
img_100.save(img_100_bytes, format="PNG")
img_100_bytes.seek(0)

# Case 5.1: Upload Document success
process = "Send POST multipart/form-data to /upload-document with image, user email, and target language."
files = {"file": ("test_form.png", img_100_bytes, "image/png")}
data = {
    "user_email": "testuser@example.com",
    "language": "Telugu"
}
res = client.post("/upload-document", data=data, files=files)
record_result("/upload-document", "POST", "Upload Document - Success uploading and OCR processing", process, res)

# Retrieve document ID from SQLite to test dependent routes
db = SessionLocal()
mock_doc = db.query(Document).filter(Document.user_email == "testuser@example.com").first()
doc_id = mock_doc.id if mock_doc else 1
db.close()

# Case 5.2: Get history
process = "Send GET request to /history/{email} to retrieve document list for a user."
res = client.get("/history/testuser@example.com")
record_result("/history/{email}", "GET", "Get History - Success listing user documents", process, res)

# Case 5.3: Get form details success
process = f"Send GET request to /form-details/{doc_id} to retrieve details of a specific uploaded form."
res = client.get(f"/form-details/{doc_id}")
record_result("/form-details/{doc_id}", "GET", "Get Form Details - Success for valid document ID", process, res)

# Case 5.4: Get form details document not found
process = "Send GET request to /form-details/9999 for a non-existent document ID."
res = client.get("/form-details/9999")
record_result("/form-details/{doc_id}", "GET", "Get Form Details - Fail for non-existent document ID", process, res)


# --- 6. FEEDBACK ENDPOINTS TESTS ---
# Case 6.1: Submit feedback (/submit-feedback)
process = "Send POST request to /submit-feedback with rating and app experience details."
payload = {
    "user_email": "testuser@example.com",
    "app_experience": "Excellent",
    "voice_guidance_helpful": "Yes",
    "recommend_app": "Yes",
    "additional_comments": "Great application!",
    "rating": "5",
    "feedback_text": "Highly intuitive design"
}
res = client.post("/submit-feedback", json=payload)
record_result("/submit-feedback", "POST", "Submit Feedback - Success submitting app feedback", process, res)

# Case 6.2: Submit feedback with slash alias (/submit-feedback/)
process = "Send POST request to /submit-feedback/ alias endpoint."
res = client.post("/submit-feedback/", json=payload)
record_result("/submit-feedback/", "POST", "Submit Feedback Slash Alias - Success", process, res)

# Case 6.3: Feedback alias endpoint (/feedback)
process = "Send POST request to /feedback alias endpoint."
res = client.post("/feedback", json=payload)
record_result("/feedback", "POST", "Submit Feedback Alias - Success", process, res)

# Case 6.4: Feedback slash alias endpoint (/feedback/)
process = "Send POST request to /feedback/ alias endpoint."
res = client.post("/feedback/", json=payload)
record_result("/feedback/", "POST", "Submit Feedback Alias Slash - Success", process, res)


# --- 7. CHANGE PASSWORD TESTS ---
# Case 7.1: Change password success
process = "Send POST request to /change-password with correct email, new password, and matching confirm password."
res = client.post("/change-password", json={
    "email": "testuser@example.com",
    "new_password": "NewStrongPassword123!",
    "confirm_password": "NewStrongPassword123!"
})
record_result("/change-password", "POST", "Change Password - Success updating registered user password", process, res)

# Case 7.2: Change password mismatch
process = "Send POST request to /change-password with mismatched passwords."
res = client.post("/change-password", json={
    "email": "testuser@example.com",
    "new_password": "NewStrongPassword123!",
    "confirm_password": "WrongPassword1!"
})
record_result("/change-password", "POST", "Change Password - Fail on mismatched password input", process, res)

# Case 7.3: Change password user not found
process = "Send POST request to /change-password with unregistered email."
res = client.post("/change-password", json={
    "email": "fake@example.com",
    "new_password": "NewStrongPassword123!",
    "confirm_password": "NewStrongPassword123!"
})
record_result("/change-password", "POST", "Change Password - Fail for non-existent user", process, res)


# --- 8. ADMIN DASHBOARD & PROTECTED ADMIN API TESTS ---
# Case 8.1: Admin login success
process = "Send POST request to /api/admin/login with correct admin email and password from configuration."
res = client.post("/api/admin/login", json={
    "email": "admin@formsahayak.com",
    "password": "AdminPassword123!"
})
record_result("/api/admin/login", "POST", "Admin Login - Success with valid administrator credentials", process, res)

# Retrieve admin access token for authenticated tests
admin_token = ""
if res.status_code == 200:
    admin_token = res.json().get("access_token", "")

admin_headers = {"Authorization": f"Bearer {admin_token}"}

# Case 8.2: Admin login invalid credentials
process = "Send POST request to /api/admin/login with incorrect credentials."
res = client.post("/api/admin/login", json={
    "email": "admin@formsahayak.com",
    "password": "WrongPassword"
})
record_result("/api/admin/login", "POST", "Admin Login - Fail on incorrect admin credentials", process, res)

# Case 8.3: Admin stats success
process = "Send GET request to /api/admin/stats with valid Bearer token in headers."
res = client.get("/api/admin/stats", headers=admin_headers)
record_result("/api/admin/stats", "GET", "Get Admin Stats - Success retrieving stats", process, res)

# Case 8.4: Admin stats unauthorized (no token)
process = "Send GET request to /api/admin/stats without any authorization headers."
res = client.get("/api/admin/stats")
record_result("/api/admin/stats", "GET", "Get Admin Stats - Fail when token is missing", process, res)

# Case 8.5: Admin users success
process = "Send GET request to /api/admin/users with valid Bearer token."
res = client.get("/api/admin/users", headers=admin_headers)
record_result("/api/admin/users", "GET", "Get Admin Users - Success listing system users", process, res)

# Case 8.6: Admin forms success
process = "Send GET request to /api/admin/forms with valid Bearer token."
res = client.get("/api/admin/forms", headers=admin_headers)
record_result("/api/admin/forms", "GET", "Get Admin Forms - Success listing uploaded forms", process, res)

# Case 8.7: Admin feedback success
process = "Send GET request to /api/admin/feedback with valid Bearer token."
res = client.get("/api/admin/feedback", headers=admin_headers)
record_result("/api/admin/feedback", "GET", "Get Admin Feedback - Success listing app feedbacks", process, res)

# Case 8.8: Serve admin dashboard UI
process = "Send GET request to /admin to load index.html file."
res = client.get("/admin")
record_result("/admin", "GET", "Serve Admin UI - Success serving HTML landing page", process, res)


# --- 9. DEVELOPER DETAILS & UPDATE DEVELOPMENT DETAILS ---
# Case 9.1: Get developer details success
process = "Send GET request to /api/developer to query developer info."
res = client.get("/api/developer")
record_result("/api/developer", "GET", "Get Developer Details - Success retrieving portfolios", process, res)

# Case 9.2: Update developer details success
process = "Send POST request to /api/admin/developer with form details and valid Bearer token."
dev_data = {
    "name": "Yamanuri Praneeth",
    "father_name": "s/o Yamanuri Ramesh",
    "role": "Developer & Creator of FormSahayak",
    "description": "Computer Science student",
    "email": "praneethyamanuri@gmail.com",
    "github": "https://github.com/Praneeth5158",
    "linkedin": "https://www.linkedin.com/in/yamanuri-praneeth/",
    "portfolio": "https://praneeth5158.github.io/My-Portfolio/"
}
res = client.post("/api/admin/developer", data=dev_data, headers=admin_headers)
record_result("/api/admin/developer", "POST", "Update Developer Details - Success updating portfolios", process, res)


# --- 10. HISTORY DELETION ENDPOINT TESTS ---
# Case 10.1: Delete history record success
process = f"Send DELETE request to /delete-history/{doc_id} to remove a specific record."
res = client.delete(f"/delete-history/{doc_id}")
record_result("/delete-history/{doc_id}", "DELETE", "Delete History - Success removing user document record", process, res)

# Case 10.2: Delete history record not found
process = "Send DELETE request to /delete-history/9999 to remove a non-existent record."
res = client.delete("/delete-history/9999")
record_result("/delete-history/{doc_id}", "DELETE", "Delete History - Fail for non-existent document ID", process, res)

# ==============================================================================
# REPORT EXCEL GENERATION
# ==============================================================================
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

print("\n--- GENERATING EXCEL REPORT ---")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "API Testing Report"

# Set sheet grid lines visible
ws.views.sheetView[0].showGridLines = True

# Colors and Fills
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Blue
title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")

pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
pass_font = Font(name="Calibri", size=10, bold=True, color="375623")

fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red
fail_font = Font(name="Calibri", size=10, bold=True, color="C00000")

regular_font = Font(name="Calibri", size=10)
italic_font = Font(name="Calibri", size=10, italic=True)

# Alignments
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Borders
thin_side = Side(style='thin', color='D9D9D9')
thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Title Block
ws.merge_cells("A1:G1")
ws["A1"] = "FormSahayak Backend API Testing Report"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 40

# Subtitle Info
ws["A2"] = "Date: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws["A2"].font = italic_font
ws["A3"] = "Total Test Cases: " + str(len(test_results))
ws["A3"].font = italic_font

pass_count = sum(1 for r in test_results if r["Status"] == "PASS")
fail_count = sum(1 for r in test_results if r["Status"] == "FAIL")

ws["C3"] = f"Passed: {pass_count} | Failed: {fail_count}"
ws["C3"].font = Font(name="Calibri", size=10, bold=True, color="375623" if fail_count == 0 else "C00000")

# Spacer row
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 12

# Table Headers
headers = ["Endpoint", "Method", "Test Case", "Testing Process", "Status", "Response Code", "Response Content"]
start_row = 5
ws.row_dimensions[start_row].height = 28

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=start_row, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Write Data
for idx, res_dict in enumerate(test_results):
    curr_row = start_row + 1 + idx
    ws.row_dimensions[curr_row].height = 24
    
    # Write values
    c_ep = ws.cell(row=curr_row, column=1, value=res_dict["Endpoint"])
    c_m = ws.cell(row=curr_row, column=2, value=res_dict["Method"])
    c_tc = ws.cell(row=curr_row, column=3, value=res_dict["Test Case"])
    c_p = ws.cell(row=curr_row, column=4, value=res_dict["Testing Process"])
    c_s = ws.cell(row=curr_row, column=5, value=res_dict["Status"])
    c_rc = ws.cell(row=curr_row, column=6, value=res_dict["Response Code"])
    c_c = ws.cell(row=curr_row, column=7, value=res_dict["Response Content"])
    
    # Styles
    for col_c in [c_ep, c_tc, c_p, c_c]:
        col_c.font = regular_font
        col_c.alignment = left_align
        col_c.border = thin_border
        
    for col_c in [c_m, c_rc]:
        col_c.font = regular_font
        col_c.alignment = center_align
        col_c.border = thin_border

    # Status column conditional formatting
    c_s.alignment = center_align
    c_s.border = thin_border
    if res_dict["Status"] == "PASS":
        c_s.fill = pass_fill
        c_s.font = pass_font
    else:
        c_s.fill = fail_fill
        c_s.font = fail_font

# Set Column Widths
column_widths = {
    "A": 28,  # Endpoint
    "B": 12,  # Method
    "C": 35,  # Test Case
    "D": 60,  # Process
    "E": 12,  # Status
    "F": 16,  # Response Code
    "G": 55   # Content
}

for col_letter, width in column_widths.items():
    ws.column_dimensions[col_letter].width = width

report_name = "API_Testing_Report.xlsx"
wb.save(report_name)

# ==============================================================================
# CLEANUP
# ==============================================================================
# Remove temporary files created during OCR and update-profile tests
for folder in ["uploads", "audio", "pdfs"]:
    if os.path.exists(folder):
        for filename in os.listdir(folder):
            if filename.startswith("profile_") or filename.startswith("dev_") or "test_form.png" in filename or filename.endswith(".pdf") or filename.endswith(".mp3"):
                file_path = os.path.join(folder, filename)
                try:
                    os.remove(file_path)
                except Exception:
                    pass

print(f"\nExcel report generated and saved as '{report_name}'")
print(f"Total Test Cases: {len(test_results)}")
print(f"Passed: {pass_count} ({pass_count/len(test_results)*100:.1f}%)")
print(f"Failed: {fail_count} ({fail_count/len(test_results)*100:.1f}%)")
print("Finished!\n")
